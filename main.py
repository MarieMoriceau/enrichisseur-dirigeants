import os, json, asyncio, httpx, re, smtplib, csv, io, sqlite3, time
from io import BytesIO
from pathlib import Path
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from fastapi.templating import Jinja2Templates
app = FastAPI()
templates = Jinja2Templates(directory="templates")


# ────────────────────────────────────────────────────────────────────
# GLOBAL EXCEPTION HANDLER — log le traceback dans les logs Render
# et renvoie un JSON exploitable côté HF Space (au lieu d'un 500 muet)
# ────────────────────────────────────────────────────────────────────
import traceback
from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    tb = traceback.format_exc()
    print(f"[FATAL ERROR] {request.url.path} | {type(exc).__name__}: {exc}")
    print(tb)
    return JSONResponse(
        status_code=500,
        content={
            "error": type(exc).__name__,
            "message": str(exc)[:500],
            "path": str(request.url.path),
            "traceback_tail": tb[-1500:],
        },
    )

ANTHROPIC_KEY  = os.getenv("ANTHROPIC_API_KEY", "")
PAPPERS_KEY    = os.getenv("PAPPERS_API_KEY", "")
FULLENRICH_KEY = os.getenv("FULLENRICH_API_KEY", "")
PIPEDRIVE_KEY  = os.getenv("PIPEDRIVE_API_KEY", "")
KASPR_KEY      = os.getenv("KASPR_API_KEY", "")

# Modèles Claude — Sonnet pour la qualité, Haiku pour les tâches simples (4× moins cher)
MODEL_SONNET = "claude-sonnet-4-6"          # enrich_one, enrich_claude (qualité critique)
MODEL_HAIKU  = "claude-haiku-4-5-20251001"  # trouver_linkedin, corriger_domaine (extraction simple)

# System prompt pour enrich_one/enrich_claude — bloc stable cachable
# (doit faire >= 1024 tokens pour activer le prompt caching Anthropic)
SYSTEM_ENRICH = """Tu es un assistant spécialisé dans la recherche de dirigeants opérationnels de sociétés françaises sur le web.

OBJECTIF
Identifier les vrais décideurs opérationnels (CEO, DG, CFO, DAF, CTO, COO, CMO, DRH, Président, Gérant, Partners, Associés, Fondateurs) — pas les simples mandataires légaux ou représentants de holdings.

CONSIGNES STRICTES
- Cherche uniquement les dirigeants ACTUELLEMENT EN POSTE.
- Ne jamais inclure : "ancien", "ex-", "sortant", "démissionnaire", "jusqu'au".
- Postes à EXCLURE absolument : commissaire aux comptes, conseil de surveillance, membre du conseil, censeur, observateur, représentant permanent, liquidateur, mandataire, administrateur judiciaire.
- Privilégie les sources officielles : site de la société (page "équipe", "à propos", "leadership"), LinkedIn, presse spécialisée (Les Échos, Maddyness, Frenchweb).
- Si la société est une filiale d'un groupe, identifie les dirigeants de l'entité française précisément ciblée (pas ceux du groupe parent).

RÈGLES EMAIL
- Emails uniquement professionnels (domaine de la société). Jamais gmail, hotmail, yahoo, outlook.com, icloud, free, orange, wanadoo.
- Si tu trouves l'email confirmé sur le site officiel ou une source fiable → "confiance_email": "haute"
- Si tu déduis l'email du pattern habituel (prenom.nom@domaine, p.nom@domaine, etc.) sans confirmation → "confiance_email": "moyenne"
- Si tu n'es pas sûr ou pattern incertain → "email": null, "confiance_email": "faible"

FORMAT DE SORTIE OBLIGATOIRE
Réponds UNIQUEMENT avec ce JSON, sans aucun texte avant ni après, sans backticks, sans markdown :
{"contacts":[{"prenom":"...","nom":"...","titre":"...","email":"...ou null","confiance_email":"haute|moyenne|faible"}]}

EXEMPLES DE BONNES RÉPONSES

Exemple 1 — startup tech avec dirigeants identifiables sur le site :
Input : Société "Doctolib", site doctolib.fr
Output : {"contacts":[{"prenom":"Stanislas","nom":"Niox-Chateau","titre":"CEO & Cofondateur","email":"stanislas.niox-chateau@doctolib.com","confiance_email":"moyenne"},{"prenom":"Olivier","nom":"Loison","titre":"COO","email":"olivier.loison@doctolib.com","confiance_email":"moyenne"}]}

Exemple 2 — société avec plusieurs cofondateurs dirigeants :
Input : Société "Mirakl", site mirakl.com
Output : {"contacts":[{"prenom":"Adrien","nom":"Nussenbaum","titre":"CEO & Cofondateur","email":"adrien.nussenbaum@mirakl.com","confiance_email":"moyenne"},{"prenom":"Philippe","nom":"Corrot","titre":"Président & Cofondateur","email":"philippe.corrot@mirakl.com","confiance_email":"moyenne"}]}

Exemple 3 — distinction ancien vs actuel dirigeant :
Si une source dit "Jean Dupont, ancien CEO jusqu'en 2023" → NE PAS l'inclure.
Si une source dit "Marie Martin, CEO depuis janvier 2024" → l'inclure.

Exemple 4 — pas de dirigeant identifiable de manière fiable :
Output : {"contacts":[]}

Exemple 5 — distinguer dirigeant opérationnel vs représentant légal :
Si Pappers a renvoyé "Holding Patrimoniale Dupont, représentant Jean Dupont, président" → cherche en plus le directeur général opérationnel sur le site/LinkedIn.

Exemple 6 — emails à exclure systématiquement :
{"email":"ceo@gmail.com"} → INCORRECT, mettre "email": null à la place.

RÈGLE FINALE
Ne renvoie JAMAIS de texte hors du JSON. Pas de "Voici les contacts trouvés", pas de commentaire, pas de markdown. Le premier caractère de ta réponse doit être { et le dernier }."""
SMTP_HOST      = os.getenv("SMTP_HOST", "pro2.mail.ovh.net")
SMTP_PORT      = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER      = os.getenv("SMTP_USER", "")
SMTP_PASS      = os.getenv("SMTP_PASS", "")
# Destinataire des alertes "solde API à 0" et "run failed"
ALERT_EMAIL    = os.getenv("ALERT_EMAIL", "mmoriceau@equation-sie.com")

# ────────────────────────────────────────────────────────────────────
# CACHE SQLite — économise les appels Claude/Pappers répétés
# Utilise /var/data si Render Disk monté, sinon /tmp (perd au redéploy)
# - Cache SIREN → contacts Claude+web (TTL 60 jours)
# - Cache domaine SIREN → domaine officiel (TTL 90j si trouvé, 1j si vide)
# ────────────────────────────────────────────────────────────────────
CACHE_DIR = "/var/data" if os.path.isdir("/var/data") else "/tmp"
Path(CACHE_DIR).mkdir(parents=True, exist_ok=True)
CACHE_DB = os.path.join(CACHE_DIR, "enrich_cache.db")
CONTACTS_TTL_DAYS = 60
DOMAINE_TTL_DAYS = 90
DOMAINE_EMPTY_TTL_DAYS = 1
# FullEnrich met sa base à jour ~mensuellement — TTL de 30j cohérent.
EFFECTIF_TTL_DAYS = 30
# Bande effectif cible (salariés). Toute société dont FullEnrich Search
# annonce un effectif CLAIREMENT hors de cette bande est court-circuitée
# avant Kaspr/FullEnrich emails (Option B, validée le 08/06/2026).
EFFECTIF_MIN_BAND = int(os.getenv("EFFECTIF_MIN_BAND", "20"))
EFFECTIF_MAX_BAND = int(os.getenv("EFFECTIF_MAX_BAND", "499"))
# Cache contact-level (Kaspr + FullEnrich emails) — clé (prenom|nom|domaine).
# Évite de repayer Kaspr ou un crédit FullEnrich pour des contacts déjà
# enrichis dans les CONTACT_ENRICH_TTL_DAYS derniers jours.
CONTACT_ENRICH_TTL_DAYS = int(os.getenv("CONTACT_ENRICH_TTL_DAYS", "90"))

def _init_cache():
    with sqlite3.connect(CACHE_DB) as conn:
        # WAL mode = écritures concurrentes sans "database is locked"
        # (indispensable avec concurrency=5 sur enrich_one)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA busy_timeout=5000")  # attendre 5s si verrou
        conn.execute("""CREATE TABLE IF NOT EXISTS contacts_cache (
            siren TEXT PRIMARY KEY, contacts_json TEXT NOT NULL, cached_at INTEGER NOT NULL
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS domaine_cache (
            siren TEXT PRIMARY KEY, domaine TEXT NOT NULL, cached_at INTEGER NOT NULL
        )""")
        # Cache effectif FullEnrich (clé = domaine, TTL 30j).
        conn.execute("""CREATE TABLE IF NOT EXISTS effectif_cache (
            domaine TEXT PRIMARY KEY,
            headcount INTEGER,
            headcount_range TEXT,
            cached_at INTEGER NOT NULL
        )""")
        # Cache contact-level — Kaspr + FullEnrich emails partagent
        # la même clé (prenom|nom|domaine, normalisée). TTL 90j par défaut.
        conn.execute("""CREATE TABLE IF NOT EXISTS contact_enrich_cache (
            contact_key TEXT PRIMARY KEY,
            email TEXT,
            phone TEXT,
            linkedin TEXT,
            confiance TEXT,
            source TEXT,
            cached_at INTEGER NOT NULL
        )""")
        # Historique des runs (= Excel finaux pour re-téléchargement)
        conn.execute("""CREATE TABLE IF NOT EXISTS runs_history (
            id TEXT PRIMARY KEY,
            created_at INTEGER NOT NULL,
            filename TEXT,
            total_contacts INTEGER,
            emails_count INTEGER,
            phones_count INTEGER,
            recipient TEXT,
            rows_json TEXT NOT NULL
        )""")
        # ── REWORK SERVEUR : état vivant des runs orchestrés côté serveur ──
        # Cette table porte l'état COMPLET d'un run (entrée, résultats,
        # statut, progrès). Elle survit à un redémarrage Render : au boot,
        # les runs EN_COURS interrompus sont remis EN_ATTENTE (cf _reload_runs_from_db).
        conn.execute("""CREATE TABLE IF NOT EXISTS runs (
            id TEXT PRIMARY KEY,
            nom TEXT,
            mode TEXT,
            phases TEXT,
            statut TEXT,
            phase_courante TEXT,
            progres_traites INTEGER,
            progres_total INTEGER,
            created_at INTEGER,
            started_at INTEGER,
            finished_at INTEGER,
            entree_json TEXT,
            resultats_json TEXT,
            emails_dest TEXT,
            erreur TEXT,
            stop_demande INTEGER,
            excel_filename TEXT
        )""")
        conn.commit()


def _sqlite_conn():
    """Connection avec WAL + busy_timeout activés à chaque ouverture
    (le pragma WAL n'est persistant qu'au niveau du fichier .db, mais
    busy_timeout doit être réappliqué)."""
    conn = sqlite3.connect(CACHE_DB, timeout=5)
    conn.execute("PRAGMA busy_timeout=5000")
    return conn
_init_cache()
print(f"[CACHE] Init OK → {CACHE_DB}")

def _normalize_siren(siren) -> str:
    return re.sub(r'\D', '', str(siren or ''))[:9]

def cache_contacts_get(siren: str):
    siren = _normalize_siren(siren)
    if not siren or len(siren) != 9: return None
    try:
        with _sqlite_conn() as conn:
            row = conn.execute("SELECT contacts_json, cached_at FROM contacts_cache WHERE siren=?", (siren,)).fetchone()
            if row and (time.time() - row[1]) < CONTACTS_TTL_DAYS * 86400:
                return json.loads(row[0])
    except Exception as e:
        print(f"[CACHE GET ERROR] {siren}: {e}")
    return None

def cache_contacts_set(siren: str, contacts: list):
    siren = _normalize_siren(siren)
    if not siren or len(siren) != 9 or not contacts: return
    try:
        with _sqlite_conn() as conn:
            conn.execute("INSERT OR REPLACE INTO contacts_cache (siren, contacts_json, cached_at) VALUES (?, ?, ?)",
                         (siren, json.dumps(contacts, ensure_ascii=False), int(time.time())))
            conn.commit()
    except Exception as e:
        print(f"[CACHE SET ERROR] {siren}: {e}")

def cache_domaine_get(siren: str):
    siren = _normalize_siren(siren)
    if not siren or len(siren) != 9: return None
    try:
        with _sqlite_conn() as conn:
            row = conn.execute("SELECT domaine, cached_at FROM domaine_cache WHERE siren=?", (siren,)).fetchone()
            if row:
                domaine, ts = row
                ttl = DOMAINE_TTL_DAYS if domaine else DOMAINE_EMPTY_TTL_DAYS
                if (time.time() - ts) < ttl * 86400:
                    return domaine
    except Exception as e:
        print(f"[DOMAINE CACHE GET ERROR] {siren}: {e}")
    return None

def cache_domaine_set(siren: str, domaine: str):
    siren = _normalize_siren(siren)
    if not siren or len(siren) != 9: return
    try:
        with _sqlite_conn() as conn:
            conn.execute("INSERT OR REPLACE INTO domaine_cache (siren, domaine, cached_at) VALUES (?, ?, ?)",
                         (siren, domaine or "", int(time.time())))
            conn.commit()
    except Exception as e:
        print(f"[DOMAINE CACHE SET ERROR] {siren}: {e}")


def cache_effectif_get(domaine: str):
    """Renvoie {'headcount':..., 'headcount_range':...} si en cache (TTL 30j),
    None sinon. Le cache stocke aussi les 'pas trouvé' (range='') pour éviter
    de rappeler FullEnrich en boucle pour les mêmes domaines."""
    d = (domaine or "").strip().lower()
    if not d:
        return None
    try:
        with _sqlite_conn() as conn:
            row = conn.execute(
                "SELECT headcount, headcount_range, cached_at FROM effectif_cache WHERE domaine=?",
                (d,)).fetchone()
            if row and (time.time() - row[2]) < EFFECTIF_TTL_DAYS * 86400:
                return {"headcount": row[0] or 0, "headcount_range": row[1] or ""}
    except Exception as e:
        print(f"[EFFECTIF CACHE GET ERROR] {d}: {e}")
    return None


def cache_effectif_set(domaine: str, headcount: int, headcount_range: str):
    d = (domaine or "").strip().lower()
    if not d:
        return
    try:
        with _sqlite_conn() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO effectif_cache "
                "(domaine, headcount, headcount_range, cached_at) VALUES (?, ?, ?, ?)",
                (d, int(headcount or 0), headcount_range or "", int(time.time())))
            conn.commit()
    except Exception as e:
        print(f"[EFFECTIF CACHE SET ERROR] {d}: {e}")


def _contact_cache_key(prenom: str, nom: str, domaine: str) -> str:
    """Clé normalisée (lowercase, sans accents, sans espaces parasites)
    pour le cache contact. Forme : 'prenom|nom|domaine'."""
    import unicodedata
    def _norm(s):
        s = (s or "").lower().strip()
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        s = re.sub(r"\s+", " ", s).strip()
        return s
    return f"{_norm(prenom)}|{_norm(nom)}|{_norm(domaine)}"


def cache_contact_enrich_get(prenom: str, nom: str, domaine: str):
    """Renvoie {email, phone, linkedin, confiance, source} si en cache
    (TTL CONTACT_ENRICH_TTL_DAYS jours), None sinon."""
    key = _contact_cache_key(prenom, nom, domaine)
    if not key.replace("|", "").strip():
        return None
    try:
        with _sqlite_conn() as conn:
            row = conn.execute(
                "SELECT email, phone, linkedin, confiance, source, cached_at "
                "FROM contact_enrich_cache WHERE contact_key=?",
                (key,)).fetchone()
            if row and (time.time() - row[5]) < CONTACT_ENRICH_TTL_DAYS * 86400:
                return {
                    "email":     row[0] or "",
                    "phone":     row[1] or "",
                    "linkedin":  row[2] or "",
                    "confiance": row[3] or "",
                    "source":    row[4] or "",
                }
    except Exception as e:
        print(f"[CONTACT CACHE GET ERROR] {key}: {e}")
    return None


def cache_contact_enrich_set(prenom: str, nom: str, domaine: str,
                              email: str = "", phone: str = "",
                              linkedin: str = "", confiance: str = "",
                              source: str = ""):
    """Sauvegarde l'enrichissement contact. Si une entrée existe déjà
    pour cette clé, on FUSIONNE (préserve email/phone/linkedin déjà
    connus, concatène la source pour garder l'historique des providers).
    Évite d'écraser un email Kaspr quand FullEnrich revient avec rien."""
    key = _contact_cache_key(prenom, nom, domaine)
    if not key.replace("|", "").strip():
        return
    if not (email or phone or linkedin):
        return  # rien d'utile à cacher
    try:
        with _sqlite_conn() as conn:
            row = conn.execute(
                "SELECT email, phone, linkedin, confiance, source "
                "FROM contact_enrich_cache WHERE contact_key=?",
                (key,)).fetchone()
            if row:
                email     = email     or row[0] or ""
                phone     = phone     or row[1] or ""
                linkedin  = linkedin  or row[2] or ""
                confiance = confiance or row[3] or ""
                old_src = row[4] or ""
                if source and old_src and source not in old_src:
                    source = f"{old_src}{source}".replace("++", "+")
                else:
                    source = source or old_src
            conn.execute(
                "INSERT OR REPLACE INTO contact_enrich_cache "
                "(contact_key, email, phone, linkedin, confiance, source, cached_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (key, email or "", phone or "", linkedin or "",
                 confiance or "", source or "", int(time.time())))
            conn.commit()
    except Exception as e:
        print(f"[CONTACT CACHE SET ERROR] {key}: {e}")


# ────────────────────────────────────────────────────────────────────
# SYSTÈME D'ALERTES EMAIL (solde API + run failed)
# 1 envoi max par heure par API (anti-spam : si 100 contacts plantent
# en Kaspr 402, tu reçois 1 seul email, pas 100)
# ────────────────────────────────────────────────────────────────────
import time as _t_alert
_alert_state = {}  # {api_name: last_sent_timestamp}
_ALERT_THROTTLE_SECONDS = 3600  # 1 heure

def _send_alert(subject: str, body: str, recipient: str = "") -> bool:
    """Envoie un email d'alerte simple (sans pièce jointe) via SMTP OVH."""
    if not SMTP_USER or not SMTP_PASS:
        print(f"[ALERT] SMTP non configuré, impossible d'envoyer : {subject}")
        return False
    dest = recipient or ALERT_EMAIL
    try:
        msg = MIMEMultipart()
        msg['From']    = SMTP_USER
        msg['To']      = dest
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, [dest], msg.as_string())
        print(f"[ALERT] ✅ '{subject}' envoyé à {dest}")
        return True
    except Exception as e:
        print(f"[ALERT ERROR] {e}")
        return False


def _alert_api_down(api_name: str, status_code: int = 0, detail: str = "") -> None:
    """Alerte 'solde API à 0' avec throttling 1h par API."""
    now = _t_alert.time()
    last = _alert_state.get(api_name, 0)
    if now - last < _ALERT_THROTTLE_SECONDS:
        return  # déjà alerté il y a moins d'1h
    _alert_state[api_name] = now
    subject = f"🚨 [Enrichisseur] Solde {api_name} épuisé"
    body = f"""Bonjour Marie,

L'API {api_name} a renvoyé une erreur indiquant un problème de crédit / quota
(status HTTP {status_code}).

Tant que tu ne recharges pas, le pipeline d'enrichissement va se dégrader
sur cette source de données.

→ Va recharger ton solde sur le dashboard {api_name}.

Détail technique :
{detail[:300]}

Heure de détection : {datetime.now().isoformat(timespec='seconds')}

(Cette alerte est limitée à 1 envoi par heure et par API pour éviter le spam.)

— Enrichisseur Dirigeants
"""
    _send_alert(subject, body)
ANCIENS_KEYWORDS = [
    "ancien", "ancienne", "ex-", "ex ", "démissionnaire",
    "jusqu'au", "jusqu au", "sortant"
]
TITRES_EXCLUS = [
    "commissaire aux comptes", "commissaire", "conseil de surveillance",
    "membre du conseil", "membre du directoire",
    "censeur", "observateur", "représentant permanent",
    "liquidateur", "mandataire", "administrateur",
]
def nettoyer_prenom(prenom: str) -> str:
    """Supprime civilités et prénoms composés Pappers : 'M Denis' → 'Denis', 'Florian, Paul' → 'Florian'."""
    if not prenom:
        return ""
    prenom = prenom.split(",")[0].strip()
    import re
    prenom = re.sub(r'^(M\.?\s+|Mme\.?\s+|Mr\.?\s+|Dr\.?\s+|Me\.?\s+)', '', prenom, flags=re.IGNORECASE).strip()
    return prenom
def est_ancien_dirigeant(titre: str) -> bool:
    return any(kw in titre.lower() for kw in ANCIENS_KEYWORDS)
def est_titre_exclu(titre: str) -> bool:
    return any(kw in titre.lower() for kw in TITRES_EXCLUS)
def domaine_valide(d: str) -> bool:
    d = d.strip()
    return bool(d) and "." in d and " " not in d and len(d) > 3
def nettoyer_domaine(url: str) -> str:
    """Extrait le domaine depuis une URL complète."""
    if not url:
        return ""
    d = url.lower().strip()
    d = d.replace("https://", "").replace("http://", "").replace("www.", "")
    d = d.split("/")[0].strip()
    return d


# ────────────────────────────────────────────────────────────────────
# GARDE-FOU EMAIL vs DOMAINE SOCIÉTÉ
# Catche les emails « parasites » :
#  - CRM périmé (ex-employeur conservé dans Pipedrive)
#  - Kaspr qui matche un mauvais LinkedIn
#  - Emails de la maison mère / d'une filiale
#  - Emails alumni d'une école
# Renvoie True si l'email semble étranger à la société.
# ────────────────────────────────────────────────────────────────────
ISP_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "hotmail.fr", "outlook.com",
    "outlook.fr", "yahoo.com", "yahoo.fr", "icloud.com", "me.com", "mac.com",
    "free.fr", "orange.fr", "wanadoo.fr", "sfr.fr", "neuf.fr", "laposte.net",
    "aliceadsl.fr", "numericable.fr", "bbox.fr", "live.fr", "live.com",
    "msn.com", "yandex.com", "protonmail.com", "proton.me", "tutanota.com",
    "gmx.fr", "gmx.com",
}


def _email_externe_a_societe(email: str, domaine_societe: str) -> bool:
    """True si l'email a un domaine différent de la société ET n'est pas
    une boîte mail grand public. False sinon (OK, ISP générique tolérée,
    ou impossible à vérifier faute de domaine société)."""
    if not email or "@" not in email or not domaine_societe:
        return False
    email_d = email.rsplit("@", 1)[1].lower().strip()
    if email_d in ISP_DOMAINS:
        return False
    soc_d = (domaine_societe or "").lower().strip()
    if soc_d.startswith("www."):
        soc_d = soc_d[4:]
    if not soc_d:
        return False
    if email_d == soc_d:
        return False
    # Sous-domaines : mail.cbre.fr vs cbre.fr → OK
    if email_d.endswith("." + soc_d) or soc_d.endswith("." + email_d):
        return False
    return True


def noms_similaires(nom_csv: str, nom_pappers: str) -> bool:
    import unicodedata
    def normaliser(s):
        s = s.lower().strip()
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        s = re.sub(r"[.\-_/]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s
    a = normaliser(nom_csv)
    b = normaliser(nom_pappers)
    if a == b:
        return True
    # Forme compacte (sans espaces) : rattrape 'ConvictionsRH' ↔ 'Convictions RH',
    # 'S2H Group' ↔ 'S2HGroup', etc. — utile parce que les tokens ≤2 chars sont
    # filtrés plus bas et 'RH', 'Co', 'S2' sont alors invisibles à l'intersection.
    a_c = a.replace(" ", "")
    b_c = b.replace(" ", "")
    if a_c == b_c:
        return True
    # Inclusion : un nom (au moins 5 chars compacts) entièrement contenu dans
    # l'autre — gère 'Acme' vs 'Acme Conseil' ou 'GreenkeLocation' vs 'Grenke'.
    if len(a_c) >= 5 and a_c in b_c:
        return True
    if len(b_c) >= 5 and b_c in a_c:
        return True
    mots_a = set(w for w in a.split() if len(w) > 2)
    mots_b = set(w for w in b.split() if len(w) > 2)
    if not mots_a:
        return True
    return len(mots_a & mots_b) > 0
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


# -------------------------------------------------------
# ROUTE AIDE — Mode d'emploi de l'enrichisseur (HTML standalone)
# Visible sur https://enrichisseur-dirigeants.onrender.com/aide
# -------------------------------------------------------
AIDE_HTML = """<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Mode d'emploi — Enrichisseur Dirigeants</title>
<style>
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
    background: #f8fafc; color: #1e293b;
    margin: 0; padding: 24px; line-height: 1.6;
  }
  .wrap { max-width: 860px; margin: 0 auto; }

  .header {
    background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%);
    color: white; padding: 28px 32px; border-radius: 12px;
    box-shadow: 0 4px 12px rgba(30,58,95,0.15); margin-bottom: 24px;
  }
  .header h1 { margin: 0 0 8px 0; font-size: 28px; font-weight: 700; }
  .header .pipe {
    font-size: 13px; opacity: 0.95; margin-top: 8px;
  }
  .header .pipe span {
    background: rgba(255,255,255,0.18); padding: 4px 10px;
    border-radius: 6px; margin: 0 3px; display: inline-block;
  }

  .card {
    background: white; border: 1px solid #e2e8f0; border-radius: 12px;
    padding: 24px 28px; margin-bottom: 16px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
  }
  .card h2 {
    color: #1e3a5f; margin: 0 0 16px 0; font-size: 20px;
    display: flex; align-items: center; gap: 8px;
  }

  ol { padding-left: 24px; margin: 0; }
  ol li { margin-bottom: 14px; }
  ol li strong { color: #1e3a5f; }
  ol li ul { margin-top: 8px; padding-left: 20px; }
  ol li ul li { margin-bottom: 6px; font-size: 14.5px; }

  .pipeline {
    display: flex; gap: 8px; margin: 20px 0; flex-wrap: wrap;
    align-items: center; justify-content: center;
  }
  .step {
    padding: 10px 16px; border-radius: 8px; font-size: 13px;
    font-weight: 600; color: white; display: inline-flex;
    align-items: center; gap: 8px;
  }
  .step .num {
    background: rgba(255,255,255,0.3);
    width: 20px; height: 20px; border-radius: 50%;
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 11px;
  }
  .arrow { color: #94a3b8; font-size: 16px; }

  .s1 { background: #2563eb; }  /* Pappers (bleu) */
  .s2 { background: #7c3aed; }  /* Claude (violet) */
  .s3 { background: #ea580c; }  /* Pipedrive (orange) */
  .s4 { background: #0891b2; }  /* Kaspr (cyan) */
  .s5 { background: #059669; }  /* FullEnrich (vert) */

  .warn {
    background: #fffbea; border: 1px solid #facc15; border-radius: 8px;
    padding: 14px 18px; margin: 16px 0; font-size: 14px;
  }
  .warn strong { color: #92400e; }

  .tip {
    background: #eff6ff; border: 1px solid #93c5fd; border-radius: 8px;
    padding: 14px 18px; margin: 16px 0; font-size: 14px;
  }
  .tip strong { color: #1e40af; }

  table { width: 100%; border-collapse: collapse; margin-top: 12px; }
  th, td { padding: 8px 12px; text-align: left; border-bottom: 1px solid #e2e8f0;
           font-size: 14px; }
  th { background: #f1f5f9; color: #1e3a5f; font-weight: 700; }
  td code { background: #f1f5f9; padding: 2px 6px; border-radius: 4px;
            font-size: 12.5px; color: #be123c; }

  .back {
    display: inline-block; margin-top: 24px; padding: 10px 18px;
    background: #2563eb; color: white; text-decoration: none;
    border-radius: 8px; font-weight: 600; font-size: 14px;
  }
  .back:hover { background: #1e40af; }
</style>
</head>
<body>
<div class="wrap">

  <div class="header">
    <h1>📖 Mode d'emploi — Enrichisseur Dirigeants</h1>
    <div class="pipe">
      <span>CSV Sociétés</span> →
      <span>Pappers</span> →
      <span>Claude+web</span> →
      <span>Pipedrive</span> →
      <span>Kaspr</span> →
      <span>FullEnrich</span> →
      <span>Excel par email</span>
    </div>
  </div>

  <div class="card">
    <h2>🚀 Comment lancer un enrichissement</h2>
    <ol>
      <li><strong>Télécharge le template Excel</strong> en cliquant sur <code>📥 Template</code> en haut à droite. Tu auras un fichier prêt à remplir avec les bonnes colonnes.</li>
      <li><strong>Remplis 1 ligne par société</strong> dans le template :
        <ul>
          <li>🟢 <code>nom</code> = obligatoire (nom commercial de la société, ex : <em>Grenke Location</em>)</li>
          <li>🟣 <code>siren</code>, <code>domaine</code>, <code>org_id</code>, <code>person_id</code>, <code>fondateurs</code>, <code>contact_prenom</code>, <code>contact_nom</code>, <code>contact_titre</code>, <code>contact_linkedin</code>, <code>code_postal</code>, <code>ville</code>, <code>adresse</code> = optionnels (mais plus tu remplis, meilleurs sont les résultats)</li>
        </ul>
      </li>
      <li><strong>Glisse ton fichier</strong> dans la zone <code>📂 Glissez votre fichier ici</code> (ou clique pour parcourir).</li>
      <li><strong>Saisis ton email</strong> dans <code>📧 Recevoir l'Excel par email</code> (vérifie aussi tes spams). Tu peux mettre plusieurs adresses séparées par des virgules.</li>
      <li><strong>Clique sur l'un des 2 modes de lancement</strong> :
        <ul>
          <li><code>⚡ Tout lancer</code> = enchaîne automatiquement les 5 étapes du pipeline (recommandé)</li>
          <li><strong>Étape par étape</strong> = lance manuellement chaque étape (utile pour debug ou pour ne faire qu'une partie)</li>
        </ul>
      </li>
      <li><strong>Reçois l'Excel par email</strong> à la fin (ou télécharge-le directement avec <code>⬇️ Excel</code>).</li>
    </ol>
  </div>

  <div class="card">
    <h2>🔄 Le pipeline en détail</h2>
    <div class="pipeline">
      <div class="step s1"><span class="num">1</span>Pappers</div>
      <span class="arrow">→</span>
      <div class="step s2"><span class="num">2</span>Claude+web</div>
      <span class="arrow">→</span>
      <div class="step s3"><span class="num">3</span>Pipedrive</div>
      <span class="arrow">→</span>
      <div class="step s4"><span class="num">4</span>Kaspr</div>
      <span class="arrow">→</span>
      <div class="step s5"><span class="num">5</span>FullEnrich</div>
    </div>

    <table>
      <thead>
        <tr><th>Étape</th><th>Ce qu'elle fait</th><th>Ce qu'elle ramène</th></tr>
      </thead>
      <tbody>
        <tr>
          <td><strong>Pappers</strong></td>
          <td>Interroge la base légale française par SIREN ou nom + ville</td>
          <td>SIREN, domaine, représentants légaux actifs (président, gérant…)</td>
        </tr>
        <tr>
          <td><strong>Claude+web</strong></td>
          <td>Recherche en ligne les dirigeants opérationnels (CEO, CFO, CTO, COO, DG, DRH, Partners…)</td>
          <td>Personnes en poste avec emails pro probables</td>
        </tr>
        <tr>
          <td><strong>Pipedrive</strong></td>
          <td>Vérifie si la société et ses contacts existent déjà dans ton CRM</td>
          <td>Email + téléphone des contacts déjà connus</td>
        </tr>
        <tr>
          <td><strong>Kaspr</strong></td>
          <td>Pour chaque dirigeant sans email, cherche son LinkedIn puis l'email pro associé</td>
          <td>URL LinkedIn + email pro vérifié</td>
        </tr>
        <tr>
          <td><strong>FullEnrich</strong></td>
          <td>Batch final pour les contacts qui restent sans email/téléphone (par domaine)</td>
          <td>Email + téléphone (waterfall multi-providers)</td>
        </tr>
      </tbody>
    </table>
  </div>

  <div class="card">
    <h2>⚠️ Bonnes pratiques</h2>
    <div class="warn">
      <strong>⏱️ Patience</strong> : un fichier de 50 sociétés prend ~15-30 min, 200 sociétés ~1-2h. Le pipeline est volontairement séquentiel pour ne pas se faire bloquer par les API en aval. <strong>Tu peux fermer l'onglet</strong>, l'email arrivera à la fin.
    </div>
    <div class="tip">
      <strong>💡 Optimise ton input</strong> : plus tu fournis d'infos en entrée (SIREN, domaine, code postal, ville), moins l'enrichisseur tâtonne et plus le résultat est précis. Si tu connais déjà un contact (Prénom + Nom + Titre), remplis-le aussi : il sera enrichi en priorité.
    </div>
    <div class="tip">
      <strong>🎯 Filtres automatiques</strong> : les anciens dirigeants (ex-CEO, démissionnaires…) et les rôles non-opérationnels (commissaire aux comptes, censeur, liquidateur, mandataire…) sont automatiquement exclus du résultat.
    </div>
    <div class="warn">
      <strong>🚫 Cas qui peuvent rater</strong> : société radiée (signalée en rouge dans l'Excel), sociétés sans aucun dirigeant identifiable (renvoyée vide), domaine introuvable (FullEnrich sera skip). Ce n'est pas un bug, juste des limites des sources de données.
    </div>
  </div>

  <div class="card">
    <h2>📊 Le résultat</h2>
    <p>L'Excel final contient <strong>1 ligne par dirigeant</strong> trouvé, avec les colonnes :</p>
    <table>
      <thead><tr><th>Colonne</th><th>Contenu</th></tr></thead>
      <tbody>
        <tr><td><strong>Organisation</strong></td><td>Nom de la société</td></tr>
        <tr><td><strong>Prénom / Nom / Titre</strong></td><td>Identité du dirigeant</td></tr>
        <tr><td><strong>Email</strong></td><td>Email pro (en bleu si trouvé)</td></tr>
        <tr><td><strong>Téléphone</strong></td><td>Mobile en priorité, sinon fixe</td></tr>
        <tr><td><strong>LinkedIn</strong></td><td>URL profil quand trouvée</td></tr>
        <tr><td><strong>Domaine</strong></td><td>Site web officiel de la société</td></tr>
        <tr><td><strong>Confiance</strong></td><td>haute / moyenne / faible (fiabilité de l'email)</td></tr>
        <tr><td><strong>Source</strong></td><td>Pappers, Claude+web, Pipedrive, Kaspr, FullEnrich (ou combinaisons)</td></tr>
        <tr><td><strong>Dans Pipedrive</strong></td><td>"oui" si la société/contact est déjà dans ton CRM</td></tr>
      </tbody>
    </table>
    <p style="margin-top:14px;font-size:14px;color:#64748b;">Les lignes sont regroupées par société (alternance de couleur), avec un filtre Excel actif sur les en-têtes pour trier/filtrer rapidement.</p>
  </div>

  <a href="/" class="back">← Retour à l'enrichisseur</a>

</div>
</body>
</html>"""


@app.get("/aide", response_class=HTMLResponse)
async def aide():
    """Mode d'emploi visuel de l'enrichisseur, accessible directement
    sur /aide ou via un lien depuis l'index.html."""
    return HTMLResponse(content=AIDE_HTML)


# -------------------------------------------------------
# ROUTE HISTORIQUE — Liste des runs passés + re-téléchargement
# Visible sur https://enrichisseur-dirigeants.onrender.com/runs
# -------------------------------------------------------
@app.get("/runs", response_class=HTMLResponse)
async def runs_list():
    """Liste les 50 derniers enrichissements terminés avec lien de re-téléchargement."""
    try:
        with _sqlite_conn() as conn:
            rows = conn.execute(
                "SELECT id, created_at, filename, total_contacts, emails_count, "
                "phones_count, recipient FROM runs_history "
                "ORDER BY created_at DESC LIMIT 50"
            ).fetchall()
    except Exception as e:
        return HTMLResponse(f"<p>Erreur lecture historique : {e}</p>", status_code=500)

    items_html = ""
    for r in rows:
        run_id, ts, fname, total, emails_n, phones_n, recipient = r
        date_str = datetime.fromtimestamp(ts).strftime("%d/%m/%Y à %Hh%M")
        pct = (100 * emails_n // max(1, total)) if total else 0
        items_html += f"""
        <div class="run-card">
          <div class="run-head">
            <div class="run-title">📊 {fname or 'enrichissement'}</div>
            <div class="run-date">{date_str}</div>
          </div>
          <div class="run-stats">
            <span class="stat"><b>{total}</b> contacts</span>
            <span class="stat email"><b>{emails_n}</b> emails ({pct}%)</span>
            <span class="stat phone"><b>{phones_n}</b> téléphones</span>
            <span class="stat recipient">📧 {recipient or '—'}</span>
          </div>
          <a class="dl-btn" href="/runs/{run_id}/download">⬇️ Télécharger l'Excel</a>
          <span class="run-id">id : {run_id}</span>
        </div>"""

    if not items_html:
        items_html = '<p class="empty">Aucun run terminé encore. Lancez un enrichissement depuis la page principale.</p>'

    html = f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8">
<title>Historique des runs — Enrichisseur</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: 'Segoe UI', system-ui, sans-serif; background: #f1f5f9; min-height: 100vh; }}
.header {{ background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%); color: white; padding: 24px 32px; }}
.header h1 {{ font-size: 22px; font-weight: 800; }}
.header p {{ font-size: 13px; opacity: 0.8; margin-top: 4px; }}
.container {{ max-width: 1100px; margin: 0 auto; padding: 20px; }}
.run-card {{
  background: white; border-radius: 10px; padding: 18px 22px; margin-bottom: 12px;
  border: 1px solid #e2e8f0;
}}
.run-head {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; }}
.run-title {{ font-size: 14px; font-weight: 700; color: #1e3a5f; }}
.run-date {{ font-size: 12px; color: #64748b; }}
.run-stats {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 12px; font-size: 12px; color: #475569; }}
.stat b {{ color: #1e3a5f; font-size: 14px; }}
.stat.email b {{ color: #2563eb; }}
.stat.phone b {{ color: #059669; }}
.stat.recipient {{ color: #64748b; }}
.dl-btn {{
  display: inline-block; padding: 8px 16px; background: #059669; color: white;
  text-decoration: none; border-radius: 6px; font-size: 13px; font-weight: 600;
}}
.dl-btn:hover {{ background: #047857; }}
.run-id {{ font-size: 10px; color: #94a3b8; margin-left: 12px; font-family: monospace; }}
.empty {{ text-align: center; color: #64748b; padding: 40px; background: white; border-radius: 10px; }}
.back-link {{ display: inline-block; margin-top: 16px; color: white; text-decoration: none; font-size: 13px; opacity: 0.8; }}
.back-link:hover {{ opacity: 1; }}
</style>
</head><body>
<div class="header">
  <h1>📋 Historique des enrichissements</h1>
  <p>Les 50 derniers runs terminés — re-téléchargeable même 30 jours plus tard</p>
  <a href="/" class="back-link">← Retour à l'enrichisseur</a>
</div>
<div class="container">{items_html}</div>
</body></html>"""
    return HTMLResponse(content=html)


@app.get("/runs/{run_id}/download")
async def runs_download(run_id: str):
    """Re-génère et sert l'Excel d'un run passé."""
    try:
        with _sqlite_conn() as conn:
            row = conn.execute(
                "SELECT rows_json, filename FROM runs_history WHERE id = ?",
                (run_id,)
            ).fetchone()
    except Exception as e:
        return JSONResponse({"error": f"DB error: {e}"}, status_code=500)

    if not row:
        return JSONResponse({"error": "Run introuvable"}, status_code=404)

    try:
        rows_data = json.loads(row[0])
        filename = row[1] or f"run_{run_id}.xlsx"
        content = generer_excel(rows_data)
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        return JSONResponse({"error": f"Generation error: {e}"}, status_code=500)


# -------------------------------------------------------
# ROUTE ALERTE — Envoi d'un email d'alerte générique
# Appelée par le HF Space quand un run plante (mais utilisable
# pour n'importe quelle notification)
# -------------------------------------------------------
@app.post("/send_alert")
async def send_alert_route(request: Request):
    """POST {"subject": "...", "body": "...", "recipient": "..." (optionnel)}
    → envoi SMTP. Si recipient absent, utilise ALERT_EMAIL."""
    data = await request.json()
    subject = data.get("subject", "[Enrichisseur] Notification")
    body = data.get("body", "")
    recipient = (data.get("recipient") or "").strip()
    ok = _send_alert(subject, body, recipient or "")
    return {"ok": ok}


@app.get("/test_alert")
async def test_alert_route():
    """Test rapide : envoie un email d'alerte à ALERT_EMAIL.
    Visite https://enrichisseur-dirigeants.onrender.com/test_alert
    pour vérifier que les alertes arrivent bien dans ta boîte."""
    ok = _send_alert(
        subject="✅ [Enrichisseur] Test d'alerte — tout fonctionne",
        body=("Bonjour Marie,\n\n"
              "Si tu lis ceci, c'est que le système d'alerte SMTP fonctionne. "
              "Tu recevras un email de cette nature quand :\n\n"
              "• Un run de prospection plante (avec le détail de l'erreur)\n"
              "• Le solde d'une de tes APIs est épuisé (Pappers, Claude, "
              "Kaspr, FullEnrich) — 1 seul email par heure et par API pour "
              "éviter le spam.\n\n"
              f"Destinataire configuré : {ALERT_EMAIL}\n\n"
              "— Enrichisseur Dirigeants"),
    )
    return {"ok": ok, "recipient": ALERT_EMAIL,
            "info": "Si ok=true, l'email est parti. Vérifie ta boîte (et les spams)."}


@app.get("/template")
async def get_template():
    """Génère et sert le template Excel enrichisseur."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Sociétés à enrichir"
    colonnes = [
        ("nom",            True,  "Nom commercial de la société",        "Grenke Location"),
        ("siren",          False, "SIREN ou SIRET (9 ou 14 chiffres)",   "428616734"),
        ("domaine",        False, "Domaine du site web",                 "grenke.fr"),
        ("org_id",         False, "ID Pipedrive de l'organisation",      "17136"),
        ("person_id",      False, "ID Pipedrive du contact (personne)",  "29481"),
        ("fondateurs",     False, "Fondateurs connus (contexte Claude)", "Jean Dupont, Marie Martin"),
        ("contact_prenom", False, "Prénom du contact principal",         "Nathalie"),
        ("contact_nom",    False, "Nom de famille du contact",           "Seyller"),
        ("contact_titre",  False, "Poste / fonction du contact",         "CFO"),
        ("contact_linkedin",False,"URL du profil LinkedIn du contact",   "https://www.linkedin.com/in/n-seyller"),
        ("code_postal",    False, "Code postal du siège",                "75008"),
        ("ville",          False, "Ville du siège",                      "Paris"),
        ("adresse",        False, "Adresse complète du siège",           "9 Rue de Lisbonne 75008 Paris"),
    ]
    thin   = Side(style='thin', color="e2e8f0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells(f'A1:{get_column_letter(len(colonnes))}1')
    c = ws['A1']
    c.value = "📋 Template Enrichisseur Dirigeants — 1 ligne par société"
    c.font  = Font(name='Arial', bold=True, size=13, color="FFFFFF")
    c.fill  = PatternFill('solid', start_color="1e3a5f")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f'A2:{get_column_letter(len(colonnes))}2')
    c = ws['A2']
    c.value = "🟢 Colonne obligatoire    🟣 Colonne optionnelle — plus vous remplissez, meilleurs sont les résultats"
    c.font  = Font(name='Arial', size=10, color="FFFFFF")
    c.fill  = PatternFill('solid', start_color="2563eb")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    for col_idx, (nom, obligatoire, desc, exemple) in enumerate(colonnes, 1):
        c = ws.cell(row=3, column=col_idx, value=nom)
        c.font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
        c.fill  = PatternFill('solid', start_color="059669" if obligatoire else "7c3aed")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[3].height = 25
    for col_idx, (nom, obligatoire, desc, exemple) in enumerate(colonnes, 1):
        c = ws.cell(row=4, column=col_idx, value=desc)
        c.font  = Font(name='Arial', italic=True, size=9, color="475569")
        c.fill  = PatternFill('solid', start_color="f8fafc")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 35
    for col_idx, (nom, obligatoire, desc, exemple) in enumerate(colonnes, 1):
        c = ws.cell(row=5, column=col_idx, value=exemple)
        c.font  = Font(name='Arial', size=9, color="94a3b8")
        c.fill  = PatternFill('solid', start_color="f1f5f9")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[5].height = 20
    for row in range(6, 26):
        for col_idx in range(1, len(colonnes)+1):
            c = ws.cell(row=row, column=col_idx, value="")
            c.fill = PatternFill('solid', start_color="ffffff" if row % 2 == 0 else "f8fafc")
            c.border = border
            c.font = Font(name='Arial', size=10)
        ws.row_dimensions[row].height = 18
    largeurs = [22, 14, 20, 12, 28, 16, 18, 18, 12, 14, 32]
    for i, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A6'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_enrichisseur.xlsx"}
    )
async def health():
    return {
        "ok": True,
        "anthropic_key": bool(ANTHROPIC_KEY),
        "pappers_key": bool(PAPPERS_KEY),
        "fullenrich_key": bool(FULLENRICH_KEY),
        "pipedrive_key": bool(PIPEDRIVE_KEY),
        "kaspr_key": bool(KASPR_KEY),
    }
async def check_pipedrive(prenom: str, nom: str) -> dict:
    if not PIPEDRIVE_KEY or not prenom or not nom:
        return {}
    terme = f"{prenom} {nom}".strip()
    try:
        async with httpx.AsyncClient(timeout=8) as c:
            r = await c.get(
                "https://api.pipedrive.com/v1/persons/search",
                params={"term": terme, "fields": "name,email,phone", "exact_match": "false", "limit": 5, "api_token": PIPEDRIVE_KEY}
            )
            if r.status_code == 200:
                items = r.json().get("data", {}).get("items", [])
                for item in items:
                    person = item.get("item", {})
                    person_name = person.get("name", "").lower()
                    if nom.lower() in person_name or prenom.lower() in person_name:
                        email = ""
                        emails = person.get("emails", [])
                        if emails:
                            email = emails[0] if isinstance(emails[0], str) else emails[0].get("value", "")
                        phone = ""
                        phones = person.get("phones", [])
                        for ph in phones:
                            val   = ph if isinstance(ph, str) else ph.get("value", "")
                            label = (ph.get("label","") if isinstance(ph, dict) else "").lower()
                            if val and ("mobile" in label or "portable" in label or "cell" in label):
                                phone = val
                                break
                        if not phone and phones:
                            ph = phones[0]
                            phone = ph if isinstance(ph, str) else ph.get("value", "")
                        # Poste Pipedrive : pas toujours dans le résultat de
                        # recherche → on le complète via /persons/{id}.
                        poste = person.get("job_title", "") or ""
                        if not poste and person.get("id"):
                            try:
                                rp = await c.get(
                                    f"https://api.pipedrive.com/v1/persons/{person.get('id')}",
                                    params={"api_token": PIPEDRIVE_KEY})
                                if rp.status_code == 200:
                                    poste = (rp.json().get("data") or {}).get("job_title", "") or ""
                            except Exception:
                                pass
                        # ── Politique 09/06/2026 (Marie) ─────────────────────
                        # Les emails Pipedrive sont trop souvent périmés (ex-
                        # employeurs traînant dans le CRM). On ne s'en sert
                        # plus du tout — Kaspr + FullEnrich cherchent du frais.
                        # Le téléphone et le poste, eux, restent fiables.
                        if phone or poste:
                            has_email = email and "@" in email
                            print(f"[PIPEDRIVE] ✅ {terme} → tél: {phone or '—'} | "
                                  f"poste: {poste or '—'} | "
                                  f"email CRM : {'présent (ignoré)' if has_email else 'absent'}")
                            return {"email": "", "phone": phone, "job_title": poste}
    except Exception as e:
        print(f"[PIPEDRIVE ERROR] {terme}: {e}")
    return {}
async def trouver_linkedin(prenom: str, nom: str, societe: str) -> str:
    """Cherche l'URL LinkedIn du dirigeant via Claude+web (max_uses:1)."""
    if not ANTHROPIC_KEY:
        return ""
    prenom = nettoyer_prenom(prenom)
    if not prenom or not nom:
        return ""
    try:
        prompt = f"""Trouve l'URL LinkedIn exacte de cette personne :
Prénom: {prenom}
Nom: {nom}
Société: {societe}
Réponds UNIQUEMENT avec l'URL complète (ex: https://www.linkedin.com/in/prenom-nom-xxxxx/)
Si tu n'es pas certain à 100%, réponds: NON"""
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01", "anthropic-beta": "web-search-2025-03-05", "content-type": "application/json"},
                json={
                    "model": MODEL_HAIKU,  # extraction simple, Haiku suffit
                    "max_tokens": 200,
                    "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 1}],
                    "messages": [{"role": "user", "content": prompt}]
                }
            )
            print(f"[LINKEDIN] Status {r.status_code} pour {prenom} {nom}")
            if r.status_code == 200:
                all_text = " ".join(b.get("text","") for b in r.json().get("content",[]) if b.get("type")=="text").strip()
                m = re.search(r'https?://(?:www\.)?linkedin\.com/in/[^\s\)\"\'\]]+', all_text)
                if m:
                    url = m.group().rstrip('/')
                    print(f"[LINKEDIN] ✅ {prenom} {nom} → {url}")
                    return url
    except Exception as e:
        print(f"[LINKEDIN ERROR] {prenom} {nom}: {e}")
    print(f"[LINKEDIN] ❌ Pas trouvé pour {prenom} {nom}")
    return ""
def _extract_kaspr_email(d: dict) -> str:
    """Extrait l'email pro de la réponse Kaspr (structure souvent imbriquée
    sous 'profile')."""
    profile = d.get("profile", d) or d  # imbriqué OU top-level
    # 1. workEmails (B2B priorité)
    work_emails = profile.get("workEmails", []) or profile.get("work_emails", [])
    if isinstance(work_emails, list) and work_emails:
        e = work_emails[0]
        if isinstance(e, str) and "@" in e:
            return e
        if isinstance(e, dict):
            val = e.get("email") or e.get("value") or ""
            if val and "@" in val:
                return val
    # 2. emails (cherche celui marqué isCurrent en priorité)
    emails = profile.get("emails", [])
    if isinstance(emails, list) and emails:
        current = next((e for e in emails if isinstance(e, dict) and e.get("isCurrent")), None)
        if current:
            val = current.get("email") or current.get("value") or ""
            if val and "@" in val:
                return val
        e = emails[0]
        if isinstance(e, str) and "@" in e:
            return e
        if isinstance(e, dict):
            val = e.get("email") or e.get("value") or ""
            if val and "@" in val:
                return val
    # 3. champs simples (starryWorkEmail = email garanti chez Kaspr)
    starry = profile.get("starryWorkEmail", "")
    if starry and isinstance(starry, str) and "@" in starry:
        return starry
    for k in ("workEmail", "work_email", "email"):
        val = profile.get(k, "")
        if val and isinstance(val, str) and "@" in val:
            return val
    return ""


# ────────────────────────────────────────────────────────────────────
# Throttle Kaspr — l'API limite le DÉBIT (HTTP 429), indépendamment des
# crédits (B2B email = illimité). On espace les appels ET, dès qu'un 429
# arrive, on met TOUTE l'activité Kaspr en pause — pas juste un contact.
# Sinon les appels parallèles continuent de saturer l'API et le 429
# ne se libère jamais.
# ────────────────────────────────────────────────────────────────────
_KASPR_LOCK        = asyncio.Lock()
_KASPR_LAST        = [0.0]   # timestamp (monotonic) du dernier appel
_KASPR_PAUSE_UNTIL = [0.0]   # pause GLOBALE jusqu'à ce timestamp
KASPR_MIN_INTERVAL = 4.0     # secondes minimum entre 2 appels (~15/min)
KASPR_PAUSE_429    = 60.0    # pause globale de base après un 429
KASPR_MAX_429      = 4       # nombre de 429 tolérés avant abandon
# Quota Kaspr épuisé (429 + x-ratelimit-remaining=0) : on coupe Kaspr jusqu'à
# ce timestamp monotonic, au lieu de réessayer en boucle avec des pauses de 180s.
_KASPR_QUOTA_KO_UNTIL = [0.0]


async def kaspr_email(prenom: str, nom: str, linkedin_url: str) -> str:
    """Récupère l'email via Kaspr avec une URL LinkedIn — B2B uniquement.
    Débit régulé + pause GLOBALE sur 429 (tous les appels Kaspr attendent)."""
    if not KASPR_KEY or not linkedin_url:
        return ""
    # Quota journalier Kaspr épuisé → on rend la main tout de suite
    # (sinon : pauses de 180s en boucle qui figent le run et le bouton Stop).
    if time.monotonic() < _KASPR_QUOTA_KO_UNTIL[0]:
        restant = int((_KASPR_QUOTA_KO_UNTIL[0] - time.monotonic()) // 60)
        print(f"[KASPR] ⛔ Limite de requêtes atteinte — appel sauté pour {prenom} {nom} (réessai dans ~{restant} min)")
        return ""
    n_429 = 0
    while True:
        # ── Régulation du débit (verrou global) ──
        async with _KASPR_LOCK:
            now = time.monotonic()
            # une pause globale est en cours (déclenchée par un 429) ?
            if _KASPR_PAUSE_UNTIL[0] > now:
                attente = _KASPR_PAUSE_UNTIL[0] - now
                print(f"[KASPR] ⏸️ Pause globale {attente:.0f}s avant {prenom} {nom}")
                await asyncio.sleep(attente)
                now = time.monotonic()
            # espacement minimal entre 2 appels
            ecoule = now - _KASPR_LAST[0]
            if ecoule < KASPR_MIN_INTERVAL:
                await asyncio.sleep(KASPR_MIN_INTERVAL - ecoule)
            _KASPR_LAST[0] = time.monotonic()
        try:
            async with httpx.AsyncClient(timeout=20) as c:
                print(f"[KASPR] Appel pour {prenom} {nom} — {linkedin_url}")
                r = await c.post(
                    "https://api.developers.kaspr.io/profile/linkedin",
                    headers={
                        "Authorization": f"Bearer {KASPR_KEY}",
                        "Content-Type": "application/json",
                        "accept-version": "v2.0"
                    },
                    json={
                        "name": f"{prenom} {nom}",
                        "id": linkedin_url,
                        "dataToGet": ["workEmail"]
                    }
                )
                print(f"[KASPR] Status {r.status_code} pour {prenom} {nom}")
                if r.status_code == 200:
                    d = r.json()
                    email = _extract_kaspr_email(d)
                    if email:
                        print(f"[KASPR] ✅ Email trouvé : {email}")
                        return email
                    # Pas d'email dans la réponse — log les clés pour debug
                    keys = list((d.get("profile") or d).keys())[:15]
                    print(f"[KASPR] ⚠️ Réponse 200 mais aucun email extrait. Clés profile : {keys}")
                    return ""
                elif r.status_code == 429:
                    n_429 += 1
                    # En-têtes utiles pour calibrer la vraie limite de Kaspr
                    retry_after = r.headers.get("Retry-After", "")
                    rl = {k: v for k, v in r.headers.items()
                          if "ratelimit" in k.lower() or "rate-limit" in k.lower()}
                    print(f"[KASPR] 429 #{n_429} — Retry-After={retry_after or 'absent'} | "
                          f"{rl or 'pas d en-tete rate-limit'}")
                    # QUOTA ÉPUISÉ (≠ simple pic de débit) : x-ratelimit-remaining=0
                    # ou Retry-After très long → on coupe Kaspr pour tout le reste
                    # du run, plutôt que de boucler sur des pauses de 180s.
                    remaining = r.headers.get("x-ratelimit-remaining", "")
                    try:
                        ra_sec = float(retry_after) if retry_after else 0.0
                    except ValueError:
                        ra_sec = 0.0
                    if remaining == "0" or ra_sec > 600:
                        coupe = min(ra_sec, 86400.0) if ra_sec else 3600.0
                        _KASPR_QUOTA_KO_UNTIL[0] = time.monotonic() + coupe
                        print(f"[KASPR] ⛔ Limite de requêtes atteinte → Kaspr coupé ~{int(coupe // 60)} min")
                        _send_alert(
                            "⏱️ [Enrichisseur] Kaspr — limite de requêtes API atteinte",
                            "Bonjour Marie,\n\nKaspr a atteint sa limite de requêtes API "
                            "pour la fenêtre en cours (~250 appels / 24 h d'après ses en-têtes).\n\n"
                            "⚠️ Ce n'est PAS un problème de crédits ni de facturation : tes "
                            "crédits B2B email restent illimités. C'est uniquement un plafond "
                            "de débit imposé par l'API Kaspr.\n\nLes enrichissements Kaspr se "
                            "mettent en pause et reprendront après remise à zéro (sous 24 h). "
                            "FullEnrich continue normalement.\n\n— Enrichisseur Dirigeants")
                        return ""
                    if n_429 >= KASPR_MAX_429:
                        print(f"[KASPR] ⛔ Abandon {prenom} {nom} : 429 persistant ({n_429}×)")
                        return ""
                    # Pause GLOBALE : tous les appels Kaspr vont attendre
                    try:
                        pause = float(retry_after) if retry_after else KASPR_PAUSE_429 * n_429
                    except ValueError:
                        pause = KASPR_PAUSE_429 * n_429
                    pause = min(pause, 180.0)
                    _KASPR_PAUSE_UNTIL[0] = time.monotonic() + pause
                    print(f"[KASPR] ⏸️ Débit dépassé → pause globale de {pause:.0f}s")
                    continue
                elif r.status_code == 402:
                    print(f"[KASPR] Plus de crédits !")
                    _alert_api_down("Kaspr", 402, r.text[:200] if hasattr(r, 'text') else "")
                    return ""
                else:
                    print(f"[KASPR] Erreur {r.status_code}: {r.text[:100]}")
                    return ""
        except Exception as e:
            print(f"[KASPR ERROR] {prenom} {nom}: {e}")
            return ""


# ────────────────────────────────────────────────────────────────────
# RECHERCHE-ENTREPRISES (data.gouv) — SOURCE PRIMAIRE des dirigeants
# API publique gratuite, lit la même base RNE que Pappers (dirigeants
# identiques, vérifié le 22/05/2026 sur DIAC). Donne aussi l'état
# administratif (actif/radié). Pappers ne sert plus que de filet de
# secours. Throttle ~0,5 s entre appels — l'API limite le débit.
# ────────────────────────────────────────────────────────────────────
_RECH_ENT_URL     = "https://recherche-entreprises.api.gouv.fr/search"
_RECH_ENT_LOCK    = asyncio.Lock()
_RECH_ENT_LAST    = [0.0]   # timestamp (monotonic) du dernier appel
RECH_ENT_INTERVAL = 0.5     # secondes minimum entre 2 appels (~2 req/s)


def _norm_nom_propre(s: str) -> str:
    """recherche-entreprises renvoie noms/prénoms en CAPITALES.
    'CHATAIN' → 'Chatain', 'JEAN-PIERRE' → 'Jean-Pierre'. Une chaîne
    déjà en casse mixte (cas Pappers) est laissée telle quelle.
    Retire le nom d'usage entre parenthèses :
    'KADOUCH-CHASSAING (KADOUCH)' → 'Kadouch-Chassaing'."""
    s = (s or "").strip()
    s = re.sub(r'\s*\([^)]*\)', '', s).strip()
    if not s:
        return ""
    return s.title() if s == s.upper() else s


def _prenom_recherche_entreprises(prenoms: str) -> str:
    """Le champ `prenoms` liste tous les prénoms d'état civil
    ('VALÉRIE FRANÇOISE MARIE'). On ne garde que le premier, en casse
    normale → 'Valérie'."""
    tokens = (prenoms or "").replace(",", " ").split()
    return _norm_nom_propre(tokens[0]) if tokens else ""


def _dirigeants_vers_contacts(res: dict) -> list:
    """Transforme le bloc `dirigeants[]` de recherche-entreprises en
    contacts, avec les MÊMES filtres que la voie Pappers : on écarte
    les personnes morales (holdings, commissaires aux comptes), les
    anciens dirigeants et les titres exclus (administrateurs, etc.)."""
    contacts = []
    for d in res.get("dirigeants", []) or []:
        if d.get("type_dirigeant") != "personne physique":
            continue
        titre = (d.get("qualite") or "").strip() or "Représentant légal"
        if est_ancien_dirigeant(titre) or est_titre_exclu(titre):
            continue
        nom_d = _norm_nom_propre(d.get("nom", ""))
        if not nom_d:
            continue
        contacts.append({
            "prenom":    _prenom_recherche_entreprises(d.get("prenoms", "")),
            "nom":       nom_d,
            "titre":     titre,
            "email":     "",
            "confiance": "",
            "source":    "recherche-entreprises",
        })
    return contacts


async def _recherche_entreprises(siren: str = "", nom: str = "",
                                 code_postal: str = "", ville: str = ""):
    """Interroge recherche-entreprises.api.gouv.fr/search.
    Recherche par SIREN si disponible (résultat fiable), sinon par nom
    (filtré sur le code postal + contrôle `noms_similaires`, départage
    par ville). Retourne le résultat brut de l'API (dict) ou None."""
    siren_norm = _normalize_siren(siren)
    q = siren_norm or (nom or "").strip()
    if not q:
        return None
    params = {"q": q, "per_page": 5, "page": 1}
    if not siren_norm and code_postal:
        cp = re.sub(r'\D', '', str(code_postal))[:5]
        if cp:
            params["code_postal"] = cp
    delays = [4, 12]
    for attempt in range(3):
        # ── Throttle global : ~0,5 s entre 2 appels (verrou partagé) ──
        async with _RECH_ENT_LOCK:
            ecoule = time.monotonic() - _RECH_ENT_LAST[0]
            if ecoule < RECH_ENT_INTERVAL:
                await asyncio.sleep(RECH_ENT_INTERVAL - ecoule)
            _RECH_ENT_LAST[0] = time.monotonic()
        try:
            async with httpx.AsyncClient(timeout=12) as c:
                r = await c.get(_RECH_ENT_URL, params=params)
            if r.status_code == 200:
                results = r.json().get("results", []) or []
                if not results:
                    print(f"[RECH-ENT] Aucun résultat pour '{q}'")
                    return None
                if siren_norm:
                    return results[0]   # recherche par SIREN → résultat fiable
                # Recherche par nom → on garde les résultats au nom similaire,
                # puis on départage par ville si elle est connue.
                similaires = [
                    res for res in results
                    if noms_similaires(nom, res.get("nom_complet")
                                       or res.get("nom_raison_sociale") or "")
                ]
                if not similaires:
                    print(f"[RECH-ENT] '{nom}' : aucun résultat au nom similaire")
                    return None
                if ville:
                    v = ville.strip().lower()
                    for res in similaires:
                        commune = ((res.get("siege") or {}).get("libelle_commune") or "")
                        if v and v in commune.lower():
                            return res
                return similaires[0]
            if r.status_code in (429, 502, 503, 504):
                if attempt < 2:
                    print(f"[RECH-ENT] {r.status_code} pour '{q}' — retry dans {delays[attempt]}s")
                    await asyncio.sleep(delays[attempt])
                    continue
                print(f"[RECH-ENT] {r.status_code} persistant pour '{q}' — abandon")
                return None
            print(f"[RECH-ENT] Status {r.status_code} pour '{q}'")
            return None
        except Exception as e:
            print(f"[RECH-ENT ERROR] '{q}': {e}")
            if attempt < 2:
                await asyncio.sleep(delays[attempt])
    return None


async def corriger_domaine(siren: str, societe: str) -> str:
    """Tente de trouver le vrai domaine via la recherche web Claude
    (voie gratuite, priorité), puis Pappers en dernier recours.
    Met en cache 90j (1j si vide) pour éviter les appels répétés."""
    # CACHE : on a déjà cherché le domaine pour ce SIREN ?
    cached = cache_domaine_get(siren)
    if cached is not None:
        if cached:
            print(f"[DOMAINE CACHE HIT] {societe} → {cached}")
        return cached

    # ── 1. VOIE GRATUITE : recherche web Claude (priorité) ──
    if ANTHROPIC_KEY:
        try:
            prompt = f"""Quel est le nom de domaine du site web officiel de cette société française : {societe} ?
Réponds UNIQUEMENT avec le domaine (ex: example.com), sans http ni www, sans aucun autre texte."""
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01", "anthropic-beta": "web-search-2025-03-05", "content-type": "application/json"},
                    json={
                        "model": MODEL_HAIKU,  # extraction de domaine, Haiku suffit
                        "max_tokens": 50,
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 1}],
                        "messages": [{"role": "user", "content": prompt}]
                    }
                )
                if r.status_code == 200:
                    all_text = " ".join(b.get("text","") for b in r.json().get("content",[]) if b.get("type")=="text").strip()
                    domaine = nettoyer_domaine(all_text.split()[0] if all_text else "")
                    if domaine_valide(domaine):
                        print(f"[DOMAINE FIX] Claude → {domaine} pour {societe}")
                        cache_domaine_set(siren, domaine)
                        return domaine
        except Exception as e:
            print(f"[DOMAINE FIX ERROR Claude] {e}")

    # ── 2. FILET DE SECOURS : Pappers (payant), seulement si Claude a échoué ──
    if PAPPERS_KEY:
        try:
            async with httpx.AsyncClient(timeout=8) as c:
                if siren:
                    r = await c.get("https://api.pappers.fr/v2/entreprise",
                        params={"api_token": PAPPERS_KEY, "siren": siren})
                    if r.status_code == 200:
                        d = r.json()
                        domaine = nettoyer_domaine(d.get("domaine_url","") or d.get("site_web",""))
                        if domaine_valide(domaine):
                            print(f"[DOMAINE FIX] Pappers SIREN (filet de secours) → {domaine} pour {societe}")
                            cache_domaine_set(siren, domaine)
                            return domaine
                r2 = await c.get("https://api.pappers.fr/v2/recherche",
                    params={"api_token": PAPPERS_KEY, "q": societe, "par_page": 1})
                if r2.status_code == 200:
                    resultats = r2.json().get("resultats", [])
                    if resultats:
                        domaine = nettoyer_domaine(resultats[0].get("domaine_url","") or resultats[0].get("site_web",""))
                        if domaine_valide(domaine):
                            print(f"[DOMAINE FIX] Pappers nom (filet de secours) → {domaine} pour {societe}")
                            cache_domaine_set(siren, domaine)
                            return domaine
        except Exception as e:
            print(f"[DOMAINE FIX ERROR Pappers] {e}")

    # Échec : on cache le vide (TTL 1j) pour éviter de re-tenter en boucle
    cache_domaine_set(siren, "")
    return ""


# ────────────────────────────────────────────────────────────────────
# FULLENRICH COMPANY SEARCH — effectif fiable par domaine
# Source distincte du /contact/enrich/bulk déjà utilisé en Passe 5.
# /api/v2/company/search renvoie headcount + headcount_range (tranche).
# 1 crédit FullEnrich par lookup ; cache 30j pour limiter la casse.
# Docs : https://docs.fullenrich.com/api/v2/company/search/post
# ────────────────────────────────────────────────────────────────────
async def _fullenrich_company_search(domaine: str) -> dict:
    """Recherche FullEnrich par domaine exact. Renvoie
    {'headcount': int, 'headcount_range': str} ou {} si rien trouvé /
    pas de clé / erreur. Cache 30j automatique (clé = domaine)."""
    d = (domaine or "").strip().lower()
    if not d or "." not in d:
        return {}
    # Cache d'abord — y compris les 'pas trouvé' (headcount_range='')
    cached = cache_effectif_get(d)
    if cached is not None:
        if cached.get("headcount_range"):
            print(f"[EFFECTIF CACHE HIT] {d} → {cached['headcount_range']}")
        return cached
    if not FULLENRICH_KEY:
        return {}
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.post(
                "https://app.fullenrich.com/api/v2/company/search",
                headers={
                    "Authorization": f"Bearer {FULLENRICH_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "limit": 1,
                    "domains": [{"value": d, "exact_match": True}],
                },
            )
            if r.status_code in (401, 402, 403):
                _alert_api_down("FullEnrich Search", r.status_code, r.text[:200])
                return {}
            if r.status_code == 429:
                print(f"[FE SEARCH] 429 pour {d} — on saute")
                return {}
            if r.status_code != 200:
                print(f"[FE SEARCH] Status {r.status_code} pour {d}: {r.text[:120]}")
                return {}
            data = r.json() or {}
            cies = data.get("companies", []) or []
            if not cies:
                # Cache vide pour éviter de rappeler à chaque scan
                cache_effectif_set(d, 0, "")
                print(f"[FE SEARCH] Aucune société pour {d}")
                return {"headcount": 0, "headcount_range": ""}
            cie = cies[0]
            hc = int(cie.get("headcount") or 0)
            hcr = (cie.get("headcount_range") or "").strip()
            cache_effectif_set(d, hc, hcr)
            print(f"[FE SEARCH] {d} → range={hcr!r} | headcount={hc} "
                  f"| credits={(data.get('metadata') or {}).get('credits')}")
            return {"headcount": hc, "headcount_range": hcr}
    except Exception as e:
        print(f"[FE SEARCH ERROR] {d}: {e}")
        return {}


def _est_hors_bande_effectif(headcount: int, range_str: str) -> bool:
    """True si l'effectif est CLAIREMENT hors de la bande
    [EFFECTIF_MIN_BAND, EFFECTIF_MAX_BAND]. En cas de doute (range qui
    chevauche la bande, donnée absente), renvoie False — on préfère ne
    pas jeter un prospect potentiellement bon. Headcount exact (>0) prime
    sur le range. Reconnaît les formats "11-50", "1001-5000", "10001+"."""
    mini, maxi = EFFECTIF_MIN_BAND, EFFECTIF_MAX_BAND
    if headcount and headcount > 0:
        return headcount < mini or headcount > maxi
    if range_str:
        try:
            if range_str.endswith("+"):
                lo = int(range_str.rstrip("+"))
                return lo > maxi
            parts = range_str.split("-")
            lo, hi = int(parts[0]), int(parts[1])
            if lo > maxi or hi < mini:
                return True
        except (ValueError, IndexError):
            return False
    return False


# -------------------------------------------------------
# ROUTE PASSE 1 : Pappers + Claude + Pipedrive
# -------------------------------------------------------
@app.post("/enrich_one")
async def enrich_one(request: Request):
    # Fine enveloppe — la logique vit dans _enrich_one_core (réutilisée par le worker).
    return await _enrich_one_core(await request.json())


async def _enrich_one_core(data: dict):
    pipedrive_org_contacts = []  # rempli plus bas si l'org est trouvée dans Pipedrive
    nom            = data.get("nom", "")
    siren          = re.sub(r'\D', '', data.get("siren", ""))[:9]
    domaine        = nettoyer_domaine(data.get("domaine", ""))
    org_id         = data.get("org_id", "")
    fondateurs     = data.get("fondateurs", "")
    contact_prenom = data.get("contact_prenom", "")
    contact_nom    = data.get("contact_nom", "")
    contact_titre  = data.get("contact_titre", "")
    contact_email  = data.get("contact_email", "")
    code_postal    = data.get("code_postal", "")
    ville          = data.get("ville", "")
    adresse        = data.get("adresse", "")   # adresse mappée → reportée dans la sortie
    print(f"[START] {nom} | domaine={domaine} | siren={siren}")
    # ── CHECK PIPEDRIVE ORGANISATION ──
    if PIPEDRIVE_KEY:
        try:
            async with httpx.AsyncClient(timeout=8) as c:
                r = await c.get(
                    "https://api.pipedrive.com/v1/organizations/search",
                    params={"term": nom, "exact_match": "false", "limit": 5, "api_token": PIPEDRIVE_KEY}
                )
                if r.status_code == 200:
                    items = r.json().get("data", {}).get("items", [])
                    for item in items:
                        org_name = item.get("item", {}).get("name", "")
                        org_id_pipe = item.get("item", {}).get("id", "")
                        if noms_similaires(nom, org_name):
                            print(f"[PIPEDRIVE ORG] ✅ '{nom}' dans Pipedrive → récupération contacts")
                            contacts_pipe = []
                            try:
                                r2 = await c.get(
                                    f"https://api.pipedrive.com/v1/organizations/{org_id_pipe}/persons",
                                    params={"api_token": PIPEDRIVE_KEY, "limit": 50}
                                )
                                if r2.status_code == 200:
                                    persons = r2.json().get("data") or []
                                    for p in persons:
                                        prenom_p = p.get("first_name", "") or ""
                                        nom_p    = p.get("last_name", "") or ""
                                        titre_p  = p.get("job_title", "") or ""
                                        emails_p = p.get("email", []) or []
                                        email_p  = ""
                                        for e in emails_p:
                                            val = e.get("value","") if isinstance(e, dict) else str(e)
                                            if val and "@" in val:
                                                email_p = val
                                                break
                                        phones_p = p.get("phone", []) or []
                                        phone_p  = ""
                                        for ph in phones_p:
                                            val   = ph.get("value","") if isinstance(ph, dict) else str(ph)
                                            label = (ph.get("label","") if isinstance(ph, dict) else "").lower()
                                            if val and ("mobile" in label or "portable" in label or "cell" in label):
                                                phone_p = val
                                                break
                                        if not phone_p:
                                            for ph in phones_p:
                                                val = ph.get("value","") if isinstance(ph, dict) else str(ph)
                                                if val:
                                                    phone_p = val
                                                    break
                                        # Politique 09/06/2026 : email Pipedrive jeté,
                                        # tél + poste + flag conservés.
                                        contacts_pipe.append({
                                            "prenom": prenom_p, "nom": nom_p,
                                            "titre": titre_p,
                                            "email": "",            # Kaspr/FullEnrich s'en chargent
                                            "phone": phone_p,
                                            "confiance": "",        # pas d'email = pas de confiance email
                                            "source": "Pipedrive",
                                            "dans_pipedrive": "oui",
                                            "siren": siren,
                                        })
                                    print(f"[PIPEDRIVE ORG] {len(contacts_pipe)} contacts récupérés")
                            except Exception as e2:
                                print(f"[PIPEDRIVE ORG CONTACTS ERROR] {e2}")
                            pipedrive_org_contacts = contacts_pipe
                            print(f"[PIPEDRIVE ORG] {len(contacts_pipe)} Pipedrive — on continue avec Pappers + Claude pour compléter")
                            break
        except Exception as e:
            print(f"[PIPEDRIVE ORG ERROR] {e}")
    # ── SOURCE PRIMAIRE DES DIRIGEANTS : recherche-entreprises (gratuit) ──
    # API publique data.gouv, même base RNE que Pappers. Pappers ne sert
    # plus que de filet de secours (cf section 6 du handoff, 22/05/2026).
    dirigeants_contacts = []
    re_data = await _recherche_entreprises(
        siren=siren, nom=nom, code_postal=code_postal, ville=ville)
    if re_data:
        siren_re = _normalize_siren(re_data.get("siren", ""))
        if siren_re:
            siren = siren_re
        # État administratif fiable → société radiée : on s'arrête là.
        if (re_data.get("etat_administratif") or "").upper() != "A":
            print(f"[RADIÉE] {nom} (recherche-entreprises) — arrêt")
            return {"results": [{"org_id":org_id,"societe":nom,"siren":siren,"domaine":domaine,"adresse":adresse,"prenom":"","nom_dg":"","titre":"⚠️ Société radiée","email":"","confiance":"","source":"recherche-entreprises"}]}
        dirigeants_contacts = _dirigeants_vers_contacts(re_data)
        print(f"[RECH-ENT] {nom} → {len(dirigeants_contacts)} dirigeant(s) actif(s) exploitable(s)")

    # ── FILET DE SECOURS : Pappers, UNIQUEMENT si recherche-entreprises n'a
    #    pas trouvé la société du tout (re_data is None). Si elle l'a trouvée
    #    mais sans dirigeant exploitable, inutile d'appeler Pappers : les deux
    #    lisent la même base RNE → on bascule directement sur la recherche
    #    web Claude. Économie maximale (arbitrage validé par Marie, 22/05/2026).
    pappers_data = None
    if re_data is None and PAPPERS_KEY:
        print(f"[PAPPERS] Filet de secours — société introuvable via recherche-entreprises : {nom}")
        if domaine_valide(domaine):
            try:
                async with httpx.AsyncClient(timeout=10) as c:
                    r = await c.get("https://api.pappers.fr/v2/entreprise",
                        params={"api_token": PAPPERS_KEY, "site_internet": domaine})
                    if r.status_code == 200:
                        pappers_data = r.json()
                        print(f"[PAPPERS] Trouvé par domaine")
                    elif r.status_code in (401, 402, 403, 429):
                        _alert_api_down("Pappers", r.status_code, r.text[:200])
            except Exception as e:
                print(f"[PAPPERS ERROR domaine] {e}")
        if not pappers_data and not siren:
            try:
                params = {"api_token": PAPPERS_KEY, "q": nom, "par_page": 1}
                if code_postal: params["code_postal"] = code_postal
                if ville:       params["ville"] = ville
                async with httpx.AsyncClient(timeout=10) as c:
                    r = await c.get("https://api.pappers.fr/v2/recherche", params=params)
                    if r.status_code == 200:
                        resultats = r.json().get("resultats", [])
                        if resultats:
                            siren = resultats[0].get("siren", "")
                            print(f"[PAPPERS] SIREN trouvé par nom : {siren}")
                    elif r.status_code in (401, 402, 403, 429):
                        _alert_api_down("Pappers", r.status_code, r.text[:200])
            except Exception as e:
                print(f"[PAPPERS ERROR nom] {e}")
        if not pappers_data and siren:
            try:
                async with httpx.AsyncClient(timeout=10) as c:
                    r = await c.get("https://api.pappers.fr/v2/entreprise",
                        params={"api_token": PAPPERS_KEY, "siren": siren})
                    if r.status_code == 200:
                        pappers_data = r.json()
                    elif r.status_code in (401, 402, 403, 429):
                        _alert_api_down("Pappers", r.status_code, r.text[:200])
            except Exception as e:
                print(f"[PAPPERS ERROR siren] {e}")
        if pappers_data:
            if not siren:
                siren = pappers_data.get("siren", "")
            if not domaine_valide(domaine):
                domaine_pappers = nettoyer_domaine(
                    pappers_data.get("domaine_url","") or pappers_data.get("site_web","")
                )
                if domaine_valide(domaine_pappers):
                    domaine = domaine_pappers
                    print(f"[PAPPERS] Domaine récupéré : {domaine}")
            nom_pappers = pappers_data.get("nom_entreprise","") or pappers_data.get("denomination","")
            if nom_pappers and not noms_similaires(nom, nom_pappers):
                print(f"[PAPPERS] ⚠️ Mauvaise société : '{nom_pappers}' pour '{nom}' — ignoré")
                pappers_data = None
                siren = ""
                domaine = nettoyer_domaine(data.get("domaine",""))
        if pappers_data:
            if pappers_data.get("entreprise_cessee") or pappers_data.get("statut_rcs","").lower() == "radié" or pappers_data.get("statut_consolide","").lower() == "radié":
                print(f"[RADIÉE] {nom} — arrêt")
                return {"results": [{"org_id":org_id,"societe":nom,"siren":siren,"domaine":domaine,"adresse":adresse,"prenom":"","nom_dg":"","titre":"⚠️ Société radiée","email":"","confiance":"","source":"Pappers"}]}
            for rep in pappers_data.get("representants", []):
                if rep.get("personne_morale"):
                    continue
                titre = rep.get("qualite","Représentant légal")
                if est_ancien_dirigeant(titre) or est_titre_exclu(titre):
                    continue
                prenom_raw = rep.get("prenom","")
                prenom_clean = nettoyer_prenom(prenom_raw)
                dirigeants_contacts.append({
                    "prenom": prenom_clean,
                    "nom":    rep.get("nom",""),
                    "titre":  titre,
                    "email":  "",
                    "confiance": "",
                    "source": "Pappers"
                })
            print(f"[PAPPERS] {len(dirigeants_contacts)} représentants actifs | domaine={domaine}")
    if contact_prenom and contact_nom:
        contact_prenom_clean = nettoyer_prenom(contact_prenom)
        deja_present = any(
            noms_similaires(contact_prenom_clean, ct.get("prenom","")) and
            noms_similaires(contact_nom, ct.get("nom",""))
            for ct in dirigeants_contacts
        )
        if not deja_present:
            dirigeants_contacts.insert(0, {
                "prenom": contact_prenom_clean,
                "nom":    contact_nom,
                "titre":  contact_titre or "Dirigeant",
                "email":  contact_email,
                "confiance": "haute" if contact_email else "",
                "source": "Fichier source"
            })
            print(f"[SOURCE] Contact pré-rempli : {contact_prenom_clean} {contact_nom} ({contact_titre})")
        else:
            print(f"[SOURCE] Contact déjà parmi les dirigeants : {contact_prenom_clean} {contact_nom} — skip")
    claude_contacts = []
    # Recherche web Claude = FALLBACK. On ne la lance que si NI Pipedrive NI
    # recherche-entreprises NI Pappers n'ont donné de dirigeant exploitable
    # (membres du directoire déjà écartés). Si une source légale a livré, on
    # économise l'appel Claude. La phase 2 « Claude+web » reste, elle, un
    # complément lancé à la demande.
    if pipedrive_org_contacts:
        print(f"[PIPEDRIVE ORG] {nom} dans le CRM → recherche Claude sautée")
        cached_claude = None
    elif dirigeants_contacts:
        print(f"[CLAUDE] {nom} — {len(dirigeants_contacts)} dirigeant(s) déjà trouvé(s) → recherche web sautée")
        cached_claude = None
    else:
        # CACHE : si on a déjà appelé Claude+web pour ce SIREN dans les 60j, on réutilise
        cached_claude = cache_contacts_get(siren) if siren else None
    if cached_claude is not None:
        claude_contacts = cached_claude
        print(f"[CACHE HIT] {nom} (SIREN {siren}) → {len(cached_claude)} contacts Claude réutilisés")
    elif ANTHROPIC_KEY and not pipedrive_org_contacts and not dirigeants_contacts:
        noms_deja = [f"{c['prenom']} {c['nom']}".strip() for c in dirigeants_contacts]
        exclusion = f"\nDirigeants déjà connus à ne PAS inclure : {', '.join(noms_deja)}" if noms_deja else ""
        contexte_fondateurs = f"\nFondateurs connus : {fondateurs}" if fondateurs else ""
        # Bloc dynamique uniquement (les instructions sont dans SYSTEM_ENRICH, cachées)
        prompt = f"""Société française à enrichir :
Nom : {nom}{chr(10)+"Site : "+domaine if domaine_valide(domaine) else ""}{chr(10)+"SIREN : "+siren if siren else ""}{contexte_fondateurs}{exclusion}"""
        delays = [10, 25, 45]
        for attempt in range(3):
            try:
                async with httpx.AsyncClient(timeout=90) as c:
                    print(f"[CLAUDE] Tentative {attempt+1}/3 pour {nom}")
                    r = await c.post(
                        "https://api.anthropic.com/v1/messages",
                        headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01", "anthropic-beta": "web-search-2025-03-05", "content-type": "application/json"},
                        json={
                            "model": MODEL_SONNET,
                            "max_tokens": 1000,
                            "system": [
                                {"type": "text", "text": SYSTEM_ENRICH,
                                 "cache_control": {"type": "ephemeral"}}
                            ],
                            "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 1}],
                            "messages": [{"role": "user", "content": prompt}]
                        }
                    )
                    print(f"[CLAUDE] Status {r.status_code} pour {nom}")
                    if r.status_code == 401:
                        _alert_api_down("Anthropic Claude", 401, r.text[:200])
                    if r.status_code in (429, 529):
                        await asyncio.sleep(delays[attempt])
                        continue
                    if r.status_code == 200:
                        all_text = " ".join(b.get("text","") for b in r.json().get("content",[]) if b.get("type")=="text")
                        m = re.search(r'\{[\s\S]*"contacts"[\s\S]*\}', all_text)
                        if m:
                            parsed = json.loads(m.group())
                            for ct in parsed.get("contacts",[]):
                                ct["source"] = "Claude+web"
                                email = ct.get("email","") or ""
                                if any(x in email for x in ["gmail","hotmail","yahoo","outlook.com"]):
                                    ct["email"] = ""
                                    ct["confiance_email"] = "faible"
                                if est_ancien_dirigeant(ct.get("titre","")) or est_titre_exclu(ct.get("titre","")):
                                    ct["_skip"] = True
                            claude_contacts = [ct for ct in parsed.get("contacts",[]) if not ct.get("_skip")]
                            print(f"[CLAUDE OK] {len(claude_contacts)} contacts pour {nom}")
                            # Sauve en cache pour 60 jours (économise les futurs scans)
                            if siren and claude_contacts:
                                cache_contacts_set(siren, claude_contacts)
                                print(f"[CACHE SET] {nom} (SIREN {siren}) → {len(claude_contacts)} contacts")
                        break
                    else:
                        print(f"[CLAUDE ERROR DETAIL] {r.status_code}: {r.text[:300]}")
                        break
            except Exception as e:
                print(f"[CLAUDE EXCEPTION] {e}")
                if attempt < 2:
                    await asyncio.sleep(delays[attempt])
    # Les emails devinés par Claude sont presque toujours faux → on ne les
    # conserve JAMAIS (vaut aussi pour les entrées de cache antérieures).
    # Claude identifie la PERSONNE ; l'email vient de Pipedrive/Kaspr/FullEnrich.
    for _ct in claude_contacts:
        _ct["email"] = ""
        _ct["confiance_email"] = ""
    if not domaine_valide(domaine) and siren:
        domaine = await corriger_domaine(siren, nom)

    # ── Garde-fou effectif (Option B, FullEnrich Search) ─────────────
    # FullEnrich a une info effectif plus fiable que les tranches INSEE.
    # Si elle dit clairement hors bande [EFFECTIF_MIN_BAND..MAX], on
    # court-circuite ici : pas de Pipedrive contact, pas de plafond,
    # et surtout pas de Kaspr/FullEnrich emails en aval.
    # Si FullEnrich ne connaît pas la société (cas fréquent pour de très
    # jeunes structures), on continue normalement avec ce qu'on a.
    effectif_range = ""
    if domaine_valide(domaine):
        eff = await _fullenrich_company_search(domaine)
        effectif_range = eff.get("headcount_range", "") or ""
        if _est_hors_bande_effectif(eff.get("headcount", 0), effectif_range):
            marqueur = effectif_range or str(eff.get("headcount", "?"))
            print(f"[HORS CIBLE] {nom} → effectif {marqueur} "
                  f"hors bande [{EFFECTIF_MIN_BAND}-{EFFECTIF_MAX_BAND}] — arrêt")
            return {"results": [{
                "org_id": org_id, "societe": nom, "siren": siren,
                "domaine": domaine, "adresse": adresse,
                "prenom": "", "nom_dg": "",
                "titre": f"⚠️ Société hors cible (effectif {marqueur})",
                "email": "", "phone": "", "linkedin": "",
                "confiance": "", "source": "FullEnrich Search",
                "dans_pipedrive": "",
                "effectif_fullenrich": effectif_range,
            }]}

    for ct in dirigeants_contacts + claude_contacts + pipedrive_org_contacts:
        if not ct.get("domaine"):
            ct["domaine"] = domaine
    if PIPEDRIVE_KEY and not pipedrive_org_contacts:
        # Politique 09/06/2026 : on lookup Pipedrive UNIQUEMENT pour le
        # téléphone et le poste, plus pour l'email (jeté à la source).
        # On reste libre de chercher l'email via Kaspr/FullEnrich derrière.
        for ct in dirigeants_contacts + claude_contacts:
            pd_data = await check_pipedrive(ct.get("prenom",""), ct.get("nom",""))
            if not pd_data:
                continue
            if pd_data.get("phone") and not ct.get("phone"):
                ct["phone"] = pd_data["phone"]
            if pd_data.get("job_title"):
                ct["titre"] = pd_data["job_title"]   # le poste du CRM prime
            # On marque que la personne est connue du CRM (info utile),
            # mais on ne récupère plus son email.
            ct["dans_pipedrive"] = "oui"
            src = ct.get("source","") or ""
            if "Pipedrive" not in src:
                ct["source"] = (src + "+Pipedrive").lstrip("+")
    # ── Plafond de contacts par société ──────────────────────────────
    # Règle métier : 4 contacts maximum par société.
    # EXCEPTION : si la société est déjà dans Pipedrive (contacts CRM
    # trouvés), on garde UNIQUEMENT ses contacts Pipedrive — sans plafond,
    # et sans y mêler Pappers/Claude.
    if pipedrive_org_contacts:
        tous_contacts = pipedrive_org_contacts
        print(f"[LIMITE] {nom} dans Pipedrive → {len(tous_contacts)} contact(s) CRM "
              f"conservés (Pappers/Claude ignorés)")
    else:
        # Priorité de conservation : contact du fichier → Pappers → Claude.
        # (Le contact du fichier est déjà inséré en tête de dirigeants_contacts.)
        total_trouve = len(dirigeants_contacts) + len(claude_contacts)
        tous_contacts = (dirigeants_contacts + claude_contacts)[:4]
        if total_trouve > 4:
            print(f"[LIMITE] {nom} → {total_trouve} contacts trouvés, plafonné à 4")
    if not tous_contacts:
        tous_contacts = [{"prenom":"","nom":"","titre":"","email":"","confiance":"","source":""}]
    # ── Garde-fou final : email étranger au domaine de la société ──
    # Catche les ex-employeurs traînant dans Pipedrive, les emails de la
    # maison mère, et les ratés Kaspr (mauvais profil LinkedIn matché).
    # On NE supprime PAS l'email (info précieuse pour Marie) — on dégrade
    # la confiance et on marque la source pour pouvoir filtrer dans l'Excel.
    for ct in tous_contacts:
        if _email_externe_a_societe(ct.get("email",""), domaine):
            ct["confiance_email"] = "faible"
            src = (ct.get("source","") or "")
            if "(email externe)" not in src:
                ct["source"] = (src + " (email externe)").strip()
            print(f"[GARDE-FOU] {nom} → {ct.get('prenom','')} {ct.get('nom','')} : "
                  f"email {ct.get('email','')} ≠ domaine {domaine}")
    results = []
    for ct in tous_contacts:
        results.append({
            "org_id":         org_id,
            "societe":        nom,
            "siren":          siren,
            "domaine":        domaine,
            "adresse":        adresse,
            "prenom":         ct.get("prenom",""),
            "nom_dg":         ct.get("nom",""),
            "titre":          ct.get("titre",""),
            "email":          ct.get("email","") or "",
            "phone":          ct.get("phone","") or "",
            "linkedin":       ct.get("linkedin",""),
            "confiance":      ct.get("confiance_email", ct.get("confiance","")),
            "source":         ct.get("source",""),
            "dans_pipedrive": ct.get("dans_pipedrive",""),
            "effectif_fullenrich": effectif_range,
        })
    print(f"[DONE] {nom} → {len(results)} contacts | domaine={domaine}")
    return {"results": results}
@app.post("/enrich_claude")
async def enrich_claude(request: Request):
    # Fine enveloppe — la logique vit dans _enrich_claude_core (réutilisée par le worker).
    return await _enrich_claude_core(await request.json())


async def _enrich_claude_core(data: dict):
    nom        = data.get("nom", "")
    siren      = data.get("siren", "")
    domaine    = nettoyer_domaine(data.get("domaine", ""))
    fondateurs = data.get("fondateurs", "")
    max_contacts = int(data.get("max_contacts", 4))
    if not ANTHROPIC_KEY:
        return {"contacts": []}
    contexte_fondateurs = f"\nFondateurs connus : {fondateurs}" if fondateurs else ""
    prompt = f"""Société française à enrichir :
Nom : {nom}{chr(10)+"Site : "+domaine if domaine_valide(domaine) else ""}{chr(10)+"SIREN : "+siren if siren else ""}{contexte_fondateurs}
Retourne au maximum {max_contacts} contact(s)."""
    delays = [10, 25, 45]
    for attempt in range(3):
        try:
            async with httpx.AsyncClient(timeout=90) as c:
                print(f"[CLAUDE PHASE2] Tentative {attempt+1}/3 pour {nom}")
                r = await c.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01", "anthropic-beta": "web-search-2025-03-05", "content-type": "application/json"},
                    json={
                        "model": MODEL_SONNET,
                        "max_tokens": 1000,
                        "system": [
                            {"type": "text", "text": SYSTEM_ENRICH,
                             "cache_control": {"type": "ephemeral"}}
                        ],
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 1}],
                        "messages": [{"role": "user", "content": prompt}]
                    }
                )
                print(f"[CLAUDE PHASE2] Status {r.status_code} pour {nom}")
                if r.status_code in (429, 529):
                    await asyncio.sleep(delays[attempt])
                    continue
                if r.status_code == 200:
                    all_text = " ".join(b.get("text","") for b in r.json().get("content",[]) if b.get("type")=="text")
                    m = re.search(r'\{[\s\S]*"contacts"[\s\S]*\}', all_text)
                    if m:
                        parsed = json.loads(m.group())
                        contacts = []
                        for ct in parsed.get("contacts",[]):
                            # Emails devinés par Claude jamais conservés (quasi tous faux).
                            ct["email"] = ""
                            ct["confiance_email"] = ""
                            if not est_ancien_dirigeant(ct.get("titre","")) and not est_titre_exclu(ct.get("titre","")):
                                contacts.append(ct)
                        print(f"[CLAUDE PHASE2 OK] {len(contacts)} contacts pour {nom}")
                        return {"contacts": contacts}
                else:
                    print(f"[CLAUDE PHASE2 ERROR] {r.status_code}: {r.text[:200]}")
                break
        except Exception as e:
            print(f"[CLAUDE PHASE2 EXCEPTION] {e}")
            if attempt < 2:
                await asyncio.sleep(delays[attempt])
    return {"contacts": []}
@app.post("/check_pipedrive")
async def check_pipedrive_route(request: Request):
    data = await request.json()
    prenom = data.get("prenom","")
    nom    = data.get("nom","")
    result = await check_pipedrive(prenom, nom)
    return {"email": result.get("email",""), "phone": result.get("phone",""),
            "job_title": result.get("job_title","")}
@app.post("/enrich_emails")
async def enrich_emails(request: Request):
    # Fine enveloppe — la logique vit dans _enrich_emails_core (réutilisée par le worker).
    return await _enrich_emails_core(await request.json())


async def _enrich_emails_core(data: dict):
    contacts = data.get("contacts", [])
    if not contacts:
        return {"emails": {}}
    # phase : 'kaspr' = Kaspr uniquement · 'fullenrich' = FullEnrich uniquement ·
    # absent/autre = les deux (compat anciens appels). Les phases 4 et 5 du run
    # passent explicitement l'une ou l'autre → elles sont enfin distinctes.
    phase = data.get("phase", "")
    faire_kaspr      = phase in ("", "kaspr")
    faire_fullenrich = phase in ("", "fullenrich")
    # Callable optionnel fourni par le worker → permet d'interrompre
    # proprement le traitement quand l'utilisateur clique sur « Stop ».
    stop_check = data.get("_stop_check")
    emails_result = {}
    kaspr_par_nom = {}  # résultats Kaspr déjà obtenus dans ce batch (anti-doublon)

    # ── Pre-pass cache contact-level (Kaspr + FullEnrich) ────────────
    # Pour chaque contact, on regarde si on a déjà enrichi cette
    # (prenom, nom, domaine) ces 90 derniers jours. Si oui, on évite
    # à la fois l'appel Kaspr ET le crédit FullEnrich plus bas.
    nb_cache_hit = 0
    for ct in contacts:
        prenom_ck = nettoyer_prenom(ct.get("prenom",""))
        nom_ck    = ct.get("nom","") or ""
        dom_ck    = ct.get("domaine","") or ""
        if not prenom_ck or not nom_ck or not domaine_valide(dom_ck):
            continue
        cached = cache_contact_enrich_get(prenom_ck, nom_ck, dom_ck)
        if not cached:
            continue
        if not (cached.get("email") or cached.get("phone")):
            continue
        idx = str(ct.get("idx",0))
        if cached.get("email"):
            ct["email"] = cached["email"]
            ct["_cache_hit_email"] = True
        if cached.get("phone"):
            ct["phone"] = cached["phone"]
        if cached.get("linkedin"):
            ct["linkedin"] = cached["linkedin"]
        # Politique : FullEnrich n'est jamais sollicité pour le téléphone
        # → un email en cache suffit pour court-circuiter Kaspr ET FullEnrich.
        emails_result[idx] = {
            "email":    cached.get("email","") or "",
            "phone":    cached.get("phone","") or "",
            "linkedin": cached.get("linkedin","") or "",
            "source":   cached.get("source","") or "",
        }
        nb_cache_hit += 1
        print(f"[CACHE CONTACT] {prenom_ck} {nom_ck} @ {dom_ck} → email réutilisé")
    if nb_cache_hit:
        print(f"[CACHE CONTACT] {nb_cache_hit}/{len(contacts)} contacts servis par le cache")

    if KASPR_KEY and faire_kaspr:
        # On cherche un email Kaspr pour TOUS les contacts. Les emails devinés
        # par Claude ne sont pas fiables : on ne s'en sert plus pour zapper Kaspr.
        for ct in contacts:
            if stop_check and stop_check():
                print("[KASPR] ⏹ Arrêt demandé — boucle Kaspr interrompue")
                break
            if time.monotonic() < _KASPR_QUOTA_KO_UNTIL[0]:
                print("[KASPR] ⛔ Limite de requêtes atteinte — boucle Kaspr de ce lot sautée")
                break
            # Cache contact-level a déjà fourni l'email → skip Kaspr
            if ct.get("_cache_hit_email"):
                continue
            prenom = nettoyer_prenom(ct.get("prenom",""))
            nom_ct = ct.get("nom","")
            societe_ct = ct.get("societe","")
            idx = str(ct.get("idx",0))
            if not prenom or not nom_ct:
                continue
            cle = f"{prenom} {nom_ct}".lower().strip()
            # Même personne déjà traitée dans ce batch → on réutilise le résultat
            if cle in kaspr_par_nom:
                cached = kaspr_par_nom[cle]
                if cached.get("email"):
                    ct["email"] = cached["email"]
                    ct["linkedin"] = cached.get("linkedin","")
                    ct["source_kaspr"] = True
                    emails_result[idx] = {"email": cached["email"],
                                          "linkedin": cached.get("linkedin",""),
                                          "source": "+dedup"}
                    print(f"[DEDUP] {cle} → email Kaspr déjà trouvé, réutilisé")
                continue
            # Si le fichier importé fournit déjà une URL LinkedIn valide, on
            # l'utilise directement → pas de recherche Claude (plus rapide,
            # moins cher, plus fiable). Levier identifié dans la passation.
            linkedin_fichier = (ct.get("linkedin", "") or "").strip()
            if "linkedin.com/in/" in linkedin_fichier.lower():
                linkedin_url = linkedin_fichier
                print(f"[KASPR] LinkedIn fourni par le fichier — {prenom} {nom_ct} → {linkedin_url}")
            else:
                print(f"[KASPR] Recherche LinkedIn pour {prenom} {nom_ct}")
                linkedin_url = await trouver_linkedin(prenom, nom_ct, societe_ct)
            resultat = {"email": "", "linkedin": linkedin_url or ""}
            if linkedin_url:
                if idx not in emails_result:
                    emails_result[idx] = {"email": "", "linkedin": linkedin_url, "source": ""}
                else:
                    emails_result[idx]["linkedin"] = linkedin_url
                email_kaspr = await kaspr_email(prenom, nom_ct, linkedin_url)
                if email_kaspr:
                    ct["email"] = email_kaspr
                    ct["linkedin"] = linkedin_url
                    ct["source_kaspr"] = True
                    emails_result[idx] = {"email": email_kaspr, "linkedin": linkedin_url, "source": "+Kaspr"}
                    resultat["email"] = email_kaspr
                    print(f"[KASPR] ✅ {prenom} {nom_ct} → {email_kaspr}")
                    # ── Cache contact-level : on garde email + linkedin
                    #    pour qu'un futur scan évite cet appel.
                    cache_contact_enrich_set(
                        prenom, nom_ct, ct.get("domaine",""),
                        email=email_kaspr, linkedin=linkedin_url,
                        source="+Kaspr",
                    )
            kaspr_par_nom[cle] = resultat
    # Phase « Kaspr uniquement », ou arrêt demandé → on s'arrête ici.
    if not faire_fullenrich or (stop_check and stop_check()):
        return {"emails": emails_result}
    domaines_corriges = {}
    for ct in contacts:
        if not domaine_valide(ct.get("domaine","")):
            societe = ct.get("societe","")
            siren = ct.get("siren","")
            if societe not in domaines_corriges:
                print(f"[DOMAINE FIX] Correction pour {societe}...")
                domaines_corriges[societe] = await corriger_domaine(siren, societe)
            if domaines_corriges[societe]:
                ct["domaine"] = domaines_corriges[societe]
    to_enrich = []
    # Map idx → (prenom, nom, domaine) pour pouvoir cacher les résultats
    # FullEnrich par contact à l'arrivée des réponses.
    idx_to_contact = {}
    for ct in contacts:
        # On envoie TOUS les contacts à FullEnrich, sauf ceux dont Kaspr a
        # déjà trouvé l'email (résultat fiable — inutile de payer un crédit).
        if ct.get("source_kaspr"):
            continue
        # Cache contact-level a déjà fourni l'email → skip FullEnrich
        # (on ne sollicite jamais FullEnrich pour le téléphone).
        if ct.get("_cache_hit_email"):
            continue
        if not ct.get("prenom") or not ct.get("nom"):
            continue
        domaine_ct = ct.get("domaine","").strip()
        if not domaine_valide(domaine_ct):
            print(f"[FULLENRICH] Domaine toujours invalide pour {ct.get('prenom')} {ct.get('nom')} — ignoré")
            continue
        prenom_clean = nettoyer_prenom(ct["prenom"])
        idx_str = str(ct.get("idx",0))
        idx_to_contact[idx_str] = {
            "prenom": prenom_clean,
            "nom":    ct["nom"],
            "domaine": domaine_ct,
        }
        to_enrich.append({
            "firstname":    prenom_clean,
            "lastname":     ct["nom"],
            "domain":       domaine_ct,
            "company_name": ct.get("societe",""),
            # FullEnrich n'est sollicité QUE pour les emails — pas de phone.
            # (Décision Marie, 08/06/2026 : les phones viennent de Pipedrive.)
            "enrich_fields": ["contact.emails"],
            "custom": {"idx": idx_str}
        })
    if not to_enrich:
        return {"emails": emails_result}
    print(f"[FULLENRICH BATCH] {len(to_enrich)} contacts envoyés")
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.post(
                "https://app.fullenrich.com/api/v1/contact/enrich/bulk",
                headers={"Authorization": f"Bearer {FULLENRICH_KEY}", "Content-Type": "application/json"},
                json={"name": "Enrichissement batch", "datas": to_enrich}
            )
            print(f"[FULLENRICH] Status lancement : {r.status_code}")
            if r.status_code not in (200, 201):
                print(f"[FULLENRICH ERROR] {r.text[:200]}")
                if r.status_code in (401, 402, 403):
                    _alert_api_down("FullEnrich", r.status_code, r.text[:200])
                return {"emails": emails_result}
            enrichment_id = r.json().get("enrichment_id") or r.json().get("id")
            if not enrichment_id:
                return {"emails": emails_result}
            print(f"[FULLENRICH] enrichment_id={enrichment_id}")
            for attempt in range(36):
                if stop_check and stop_check():
                    print("[FULLENRICH] ⏹ Arrêt demandé — attente du résultat interrompue")
                    return {"emails": emails_result}
                await asyncio.sleep(5)
                r2 = await c.get(
                    f"https://app.fullenrich.com/api/v1/contact/enrich/bulk/{enrichment_id}",
                    headers={"Authorization": f"Bearer {FULLENRICH_KEY}"}
                )
                if r2.status_code != 200:
                    continue
                result = r2.json()
                status = result.get("status","")
                print(f"[FULLENRICH] Polling {attempt+1}/36 — status={status}")
                if status == "FINISHED":
                    emails_par_idx = {}
                    for ct_result in result.get("datas",[]):
                        idx = ct_result.get("custom",{}).get("idx","-1")
                        contact_data = ct_result.get("contact",{})
                        email_val = ""
                        for e in contact_data.get("emails",[]):
                            val = e.get("value") or e.get("email") or ""
                            if val and "@" in val:
                                email_val = val
                                break
                        phone_val = ""
                        phones = contact_data.get("phones",[])
                        for p in phones:
                            val = p.get("number") or p.get("value") or p.get("phone") or ""
                            if val:
                                phone_val = val
                                break
                        if email_val or phone_val:
                            emails_par_idx[idx] = {"email": email_val, "phone": phone_val}
                            # ── Cache contact-level : on persiste pour
                            #    qu'un futur scan évite ce crédit FullEnrich.
                            ctx = idx_to_contact.get(idx, {})
                            if ctx.get("prenom") and ctx.get("nom") and ctx.get("domaine"):
                                # Politique : seul l'email FullEnrich est mis
                                # en cache. Les phones ne sont jamais récupérés
                                # via FullEnrich (cf enrich_fields plus haut).
                                cache_contact_enrich_set(
                                    ctx["prenom"], ctx["nom"], ctx["domaine"],
                                    email=email_val,
                                    source="+Fullenrich",
                                )
                    print(f"[FULLENRICH] {len(emails_par_idx)} contacts enrichis")
                    for k, v in emails_par_idx.items():
                        if k not in emails_result:
                            emails_result[k] = {"email": v["email"], "phone": v["phone"], "source": "+Fullenrich"}
                        else:
                            if v["email"] and not emails_result[k].get("email"):
                                emails_result[k]["email"] = v["email"]
                            if v["phone"] and not emails_result[k].get("phone"):
                                emails_result[k]["phone"] = v["phone"]
                    return {"emails": emails_result}
            print(f"[FULLENRICH] Timeout 180s")
            return {"emails": emails_result}
    except Exception as e:
        print(f"[FULLENRICH EXCEPTION] {e}")
        return {"emails": emails_result}
def generer_excel(rows: list, rayon=None, effectif: str = "") -> bytes:
    # ── Fiabilité des emails ───────────────────────────────────────
    # Les emails seulement devinés par Claude sont trop souvent faux.
    # On ne conserve un email que s'il a été confirmé par une source
    # vérifiée : Kaspr, Fullenrich ou Pipedrive (CRM). Sinon on le vide
    # (mieux vaut une case vide qu'un faux email → bounce / réputation).
    SOURCES_FIABLES = ("kaspr", "dedup", "fullenrich", "pipedrive")
    nb_vides = 0
    for r in rows:
        email = (r.get("email") or "").strip()
        src   = (r.get("source") or "").lower()
        if email and not any(s in src for s in SOURCES_FIABLES):
            r["email"]     = ""
            r["confiance"] = ""
            if "non vérifié" not in (r.get("source") or ""):
                r["source"] = (r.get("source") or "") + " (non vérifié)"
            nb_vides += 1
    if nb_vides:
        print(f"[EXCEL] {nb_vides} email(s) non vérifié(s) vidé(s) — devinés par Claude")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Dirigeants enrichis"
    headers  = ['Organisation','Adresse','Prénom','Nom','Titre','Email','Téléphone','LinkedIn','Domaine','Effectif (FullEnrich)','Confiance','Source','Dans Pipedrive']
    col_map  = ['societe','adresse','prenom','nom_dg','titre','email','phone','linkedin','domaine','effectif_fullenrich','confiance','source','dans_pipedrive']
    thin     = Side(style='thin', color="e2e8f0")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    c = ws['A1']
    _titre = "Enrichissement Dirigeants"
    _crit = []
    if rayon:
        _crit.append(f"rayon {rayon} m")
    if effectif:
        _crit.append(f"effectif {effectif}")
    if _crit:
        _titre += " — recherche : " + "  ·  ".join(_crit)
    c.value = _titre
    c.font  = Font(name='Arial', bold=True, size=14, color="FFFFFF")
    c.fill  = PatternFill('solid', start_color="1e3a5f")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32
    emails_count = len([r for r in rows if r.get('email')])
    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    c = ws['A2']
    c.value = f"{len(rows)} contacts  |  {emails_count} emails trouvés  |  {len(rows)-emails_count} sans email"
    c.font  = Font(name='Arial', size=10, color="FFFFFF")
    c.fill  = PatternFill('solid', start_color="2563eb")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill('solid', start_color="2563eb")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[3].height = 28
    src_colors = {'Pappers':'eff6ff','Claude':'f5f3ff','Pipedrive':'fef3c7','Kaspr':'e0f2fe','Fullenrich':'dcfce7'}
    rows = sorted(rows, key=lambda r: (r.get('societe','') or '').lower())
    org_list = []
    for r in rows:
        s = r.get('societe','')
        if s not in org_list:
            org_list.append(s)
    org_colors = {org: ("f0f7ff" if i % 2 == 0 else "FFFFFF") for i, org in enumerate(org_list)}
    for row_idx, row in enumerate(rows, 4):
        bg = org_colors.get(row.get('societe',''), "FFFFFF")
        ws.row_dimensions[row_idx].height = 18
        for col_idx, key in enumerate(col_map, 1):
            val = str(row.get(key, '') or '')
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(vertical='center')
            c.border = border
            if key == 'email' and val:
                c.font = Font(name='Arial', size=9, color="2563eb", bold=True)
                c.fill = PatternFill('solid', start_color=bg)
            elif key == 'confiance':
                fills = {'haute':('dcfce7','166534'),'moyenne':('fef9c3','854d0e'),'faible':('fee2e2','991b1b')}
                if val in fills:
                    c.fill = PatternFill('solid', start_color=fills[val][0])
                    c.font = Font(name='Arial', size=9, color=fills[val][1], bold=True)
                else:
                    c.fill = PatternFill('solid', start_color=bg)
            elif key == 'source' and val:
                color = next((v for k,v in src_colors.items() if k in val), bg)
                c.fill = PatternFill('solid', start_color=color)
                c.font = Font(name='Arial', size=9, bold=True)
            elif key == 'dans_pipedrive' and val:
                c.fill = PatternFill('solid', start_color="fef3c7")
                c.font = Font(name='Arial', size=9, color="92400e", bold=True)
            else:
                c.fill = PatternFill('solid', start_color=bg)
    for i, w in enumerate([22,34,14,18,28,32,16,14,20,12,22,28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(rows)+3}"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
@app.post("/export_excel")
async def export_excel(request: Request):
    data = await request.json()
    rows = data.get("rows", [])
    if not rows:
        return {"ok": False}
    content = generer_excel(rows)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enrichissement_dirigeants.xlsx"}
    )
@app.post("/send_csv")
async def send_csv(request: Request):
    data = await request.json()
    emails_raw  = data.get("emails", []) or ([data.get("email")] if data.get("email") else [])
    emails_dest = [e.strip() for e in emails_raw if e and "@" in e]
    rows        = data.get("rows", [])
    filename    = data.get("filename", "enrichissement_dirigeants.xlsx")
    rayon       = data.get("rayon")
    effectif    = data.get("effectif", "")
    if not emails_dest or not rows:
        return {"ok": False, "error": "Email(s) ou données manquants"}
    if not SMTP_USER or not SMTP_PASS:
        return {"ok": False, "error": "SMTP non configuré"}
    try:
        excel_content = generer_excel(rows, rayon=rayon, effectif=effectif)
        msg = MIMEMultipart()
        msg['From']    = SMTP_USER
        msg['To']      = ", ".join(emails_dest)
        msg['Subject'] = f"Enrichissement dirigeants — {len(rows)} contacts"
        emails_count = len([r for r in rows if r.get('email')])
        body = f"""Bonjour,
Votre enrichissement est terminé.
{len(rows)} contacts exportés dont {emails_count} emails trouvés.
Fichier Excel en pièce jointe.
Enrichisseur Dirigeants"""
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(excel_content)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, emails_dest, msg.as_string())
        print(f"[EMAIL] ✅ Excel envoyé à {', '.join(emails_dest)}")

        # ─── HISTORIQUE : sauvegarde le run pour re-téléchargement futur ──
        try:
            import uuid
            run_id = uuid.uuid4().hex[:12]
            nb_phones = sum(1 for r in rows if r.get('phone'))
            with _sqlite_conn() as conn:
                conn.execute(
                    "INSERT INTO runs_history (id, created_at, filename, "
                    "total_contacts, emails_count, phones_count, recipient, rows_json) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (run_id, int(time.time()), filename, len(rows), emails_count,
                     nb_phones, ", ".join(emails_dest),
                     json.dumps(rows, ensure_ascii=False))
                )
                conn.commit()
            print(f"[RUNS] ✅ Run sauvegardé : {run_id} ({len(rows)} contacts)")
        except Exception as e:
            print(f"[RUNS SAVE ERROR] {e}")

        return {"ok": True}
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")
        return {"ok": False, "error": str(e)}


# ════════════════════════════════════════════════════════════════════
# ★ REWORK SERVEUR — orchestration des runs côté serveur
# ════════════════════════════════════════════════════════════════════
# AVANT : tout le run vivait dans le JavaScript de index.html. Le serveur
#         ne savait pas qu'un run existait → 2 runs simultanés possibles,
#         et fermer l'onglet tuait le run en cours.
# APRÈS : un worker asyncio UNIQUE exécute les runs un par un.
#   _RUNS         : registre en mémoire {run_id: état}   (source live)
#   table `runs`  : persistance SQLite (survit au redémarrage Render)
#   _run_worker   : boucle — prend le plus ancien EN_ATTENTE, l'exécute
#
#   « Jamais 2 à la fois » = garanti par le worker unique (séquentiel).
#   « File d'attente »     = les runs EN_ATTENTE triés par created_at.
#   « Survit à l'onglet »  = le run vit dans le process serveur, pas le JS.
# ════════════════════════════════════════════════════════════════════
import uuid as _uuid

_RUNS = {}   # run_id -> dict (état live d'un run)

# Délais repris à l'identique du JS d'origine (index.html) pour ne pas
# se faire bloquer par les API en aval.
_DELAI_SOCIETE = 8.0   # pause entre 2 sociétés (phases 1 & 2)
_BATCH_SIZE    = 5     # contacts par lot (phases 4 & 5) — évite le timeout proxy Render
_PAUSE_BATCH   = 5.0   # pause entre 2 lots d'emails


# ─── Persistance SQLite ─────────────────────────────────────────────
def _run_to_db_tuple(r: dict):
    return (
        r["id"], r["nom"], r["mode"], json.dumps(r["phases"]),
        r["statut"], r["phase_courante"], r["progres_traites"],
        r["progres_total"], r["created_at"], r["started_at"],
        r["finished_at"], json.dumps(r["entree"], ensure_ascii=False),
        json.dumps(r["resultats"], ensure_ascii=False),
        json.dumps(r["emails_dest"], ensure_ascii=False),
        r["erreur"], 1 if r["stop_demande"] else 0, r["excel_filename"],
    )


def _persist_run(run_id: str):
    """Écrit l'état complet du run dans la table `runs`. Appelé à chaque
    changement de phase / progrès → le run survit à un redémarrage."""
    r = _RUNS.get(run_id)
    if not r:
        return
    try:
        with _sqlite_conn() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO runs (id, nom, mode, phases, statut, "
                "phase_courante, progres_traites, progres_total, created_at, "
                "started_at, finished_at, entree_json, resultats_json, "
                "emails_dest, erreur, stop_demande, excel_filename) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                _run_to_db_tuple(r))
            conn.commit()
    except Exception as e:
        print(f"[RUNS PERSIST ERROR] {run_id}: {e}")


def _reload_runs_from_db():
    """Recharge les runs depuis SQLite au démarrage. Les runs EN_COURS
    interrompus par un redémarrage Render sont remis EN_ATTENTE (ils
    seront ré-exécutés depuis le début par le worker)."""
    try:
        with _sqlite_conn() as conn:
            rows = conn.execute(
                "SELECT id, nom, mode, phases, statut, phase_courante, "
                "progres_traites, progres_total, created_at, started_at, "
                "finished_at, entree_json, resultats_json, emails_dest, "
                "erreur, stop_demande, excel_filename FROM runs").fetchall()
    except Exception as e:
        print(f"[RUNS RELOAD ERROR] {e}")
        return
    repris = 0
    for row in rows:
        (rid, nom, mode, phases, statut, phase_c, p_tr, p_tot, c_at, s_at,
         f_at, entree_j, res_j, emails_j, erreur, stop_d, xls) = row
        if statut == "EN_COURS":
            if stop_d:
                # Un arrêt avait été demandé avant le redémarrage → on le respecte.
                statut, phase_c = "ARRETE", "Arrêté (interrompu par un redémarrage)"
            else:
                statut, phase_c, p_tr, erreur = "EN_ATTENTE", "", 0, ""
                repris += 1
        _RUNS[rid] = {
            "id": rid, "nom": nom or "", "mode": mode or "societes",
            "phases": json.loads(phases) if phases else [1, 2, 3, 4, 5],
            "statut": statut or "EN_ATTENTE", "phase_courante": phase_c or "",
            "progres_traites": p_tr or 0, "progres_total": p_tot or 0,
            "created_at": c_at or int(time.time()),
            "started_at": s_at, "finished_at": f_at,
            "entree": json.loads(entree_j) if entree_j else [],
            "resultats": json.loads(res_j) if res_j else [],
            "emails_dest": json.loads(emails_j) if emails_j else [],
            "erreur": erreur or "", "stop_demande": bool(stop_d),
            "excel_filename": xls or "",
        }
    print(f"[RUNS] Rechargés : {len(_RUNS)} run(s) — {repris} remis EN_ATTENTE après redémarrage")


# ─── File d'attente (dérivée de _RUNS, pas de structure séparée) ────
def _next_queued_run():
    """Le plus ancien run EN_ATTENTE = tête de file. None si file vide."""
    attente = [r for r in _RUNS.values() if r["statut"] == "EN_ATTENTE"]
    if not attente:
        return None
    attente.sort(key=lambda r: r["created_at"])
    return attente[0]


def _run_en_cours():
    for r in _RUNS.values():
        if r["statut"] == "EN_COURS":
            return r
    return None


def _rang_file(run: dict):
    """Position du run dans la file d'attente (1 = prochain). None si pas EN_ATTENTE."""
    if run["statut"] != "EN_ATTENTE":
        return None
    attente = sorted([r for r in _RUNS.values() if r["statut"] == "EN_ATTENTE"],
                     key=lambda r: r["created_at"])
    return attente.index(run) + 1


# ─── Worker unique ──────────────────────────────────────────────────
async def _run_worker():
    """Boucle infinie : tant qu'aucun run EN_COURS et la file non vide,
    prend le plus ancien et l'exécute jusqu'au bout. Un seul worker =
    jamais 2 runs en parallèle."""
    print("[WORKER] Démarré — en attente de runs")
    while True:
        try:
            if _run_en_cours():            # ceinture + bretelles
                await asyncio.sleep(2)
                continue
            run = _next_queued_run()
            if run is None:
                await asyncio.sleep(2)
                continue
            await _executer_run(run["id"])
        except Exception as e:
            print(f"[WORKER ERROR] {type(e).__name__}: {e}")
            print(traceback.format_exc())
            await asyncio.sleep(3)


@app.on_event("startup")
async def _startup_runs():
    """Au démarrage Render : recharge les runs persistés, lance le worker."""
    _reload_runs_from_db()
    asyncio.create_task(_run_worker())
    print("[STARTUP] Worker de runs lancé")


# ════════════════════════════════════════════════════════════════════
# ORCHESTRATEUR — rejoue les phases 1→5 du JS, côté serveur
# ════════════════════════════════════════════════════════════════════
async def _executer_run(run_id: str):
    """Exécute un run de bout en bout : phases demandées, puis Excel +
    email + archivage. Met à jour l'état à chaque étape (persisté)."""
    r = _RUNS.get(run_id)
    if not r:
        return
    r["statut"] = "EN_COURS"
    r["started_at"] = int(time.time())
    r["finished_at"] = None
    r["erreur"] = ""
    _persist_run(run_id)
    print(f"[RUN {run_id}] ▶ Démarrage — {r['nom']} | mode={r['mode']} | phases={r['phases']}")
    try:
        # Mode « contacts importés » : l'entrée EST la liste de contacts
        # → elle devient directement la base de résultats (pas de Pappers).
        if r["mode"] == "contacts" and not r["resultats"]:
            r["resultats"] = [dict(ct) for ct in r["entree"]]

        for phase in r["phases"]:
            if r["stop_demande"]:
                break
            if phase == 1:
                await _run_phase1(r)
            elif phase == 2:
                await _run_phase2(r)
            elif phase == 3:
                await _run_phase3(r)
            elif phase == 4:
                await _run_phase_emails(r, "kaspr")
            elif phase == 5:
                await _run_phase_emails(r, "fullenrich")

        if r["stop_demande"]:
            r["statut"] = "ARRETE"
            r["phase_courante"] = "Arrêté par l'utilisateur"
            r["finished_at"] = int(time.time())
            _persist_run(run_id)
            print(f"[RUN {run_id}] ⏹ Arrêté par l'utilisateur")
            return

        # ── Finalisation : Excel + email + archivage /runs ──
        r["phase_courante"] = "Finalisation (Excel + email)"
        _persist_run(run_id)
        await _finaliser_run(r)
        r["statut"] = "TERMINE"
        r["phase_courante"] = "Terminé"
        r["finished_at"] = int(time.time())
        _persist_run(run_id)
        print(f"[RUN {run_id}] ✅ Terminé — {len(r['resultats'])} contacts")
    except Exception as e:
        r["statut"] = "ERREUR"
        r["erreur"] = f"{type(e).__name__}: {e}"
        r["finished_at"] = int(time.time())
        _persist_run(run_id)
        print(f"[RUN {run_id}] ❌ ERREUR : {e}")
        print(traceback.format_exc())
        try:
            _send_alert(
                f"🚨 [Enrichisseur] Run en échec — {r['nom']}",
                f"Bonjour Marie,\n\nLe run « {r['nom']} » (id {run_id}) a planté.\n\n"
                f"Erreur : {r['erreur']}\n\n"
                f"Détail technique :\n{traceback.format_exc()[-1200:]}\n\n"
                f"— Enrichisseur Dirigeants")
        except Exception:
            pass


# ─── PHASE 1 : Pappers + Claude + Pipedrive (1 appel par société) ───
async def _run_phase1(r: dict):
    r["phase_courante"] = "1. Pappers + Claude"
    rows = r["entree"]
    r["progres_total"] = len(rows)
    r["progres_traites"] = 0
    r["resultats"] = []      # phase 1 reconstruit la base de résultats
    _persist_run(r["id"])
    for i, row in enumerate(rows):
        if r["stop_demande"]:
            return
        try:
            data = await _enrich_one_core(dict(row))
            r["resultats"].extend(data.get("results", []))
        except Exception as e:
            print(f"[RUN {r['id']}] phase1 erreur sur {row.get('nom','?')}: {e}")
            r["resultats"].append({
                "societe": row.get("nom", ""), "siren": "", "prenom": "",
                "nom_dg": "", "titre": "", "email": "", "confiance": "", "source": ""})
        r["progres_traites"] = i + 1
        _persist_run(r["id"])
        if i < len(rows) - 1 and not r["stop_demande"]:
            await asyncio.sleep(_DELAI_SOCIETE)


# ─── PHASE 2 : Claude+web — complète les sociétés < 3 contacts ──────
async def _run_phase2(r: dict):
    r["phase_courante"] = "2. Claude + web"
    rows = r["entree"]
    resultats = r["resultats"]
    societes = []
    for x in resultats:
        s = x.get("societe", "")
        if s and s not in societes:
            societes.append(s)
    r["progres_total"] = len(societes)
    r["progres_traites"] = 0
    _persist_run(r["id"])
    for i, nom_soc in enumerate(societes):
        if r["stop_demande"]:
            return
        existants = [x for x in resultats if x.get("societe") == nom_soc]
        # Skip si déjà récupérée depuis Pipedrive en phase 1
        if any(x.get("dans_pipedrive") for x in existants):
            r["progres_traites"] = i + 1
            continue
        contacts_reels = [x for x in existants if x.get("nom_dg")]
        if len(contacts_reels) >= 4:
            r["progres_traites"] = i + 1
            continue
        max_ajouter = 4 - len(contacts_reels)
        row = dict(next((rw for rw in rows if rw.get("nom") == nom_soc),
                        {"nom": nom_soc, "domaine": "", "siren": "", "fondateurs": ""}))
        ex = next((x for x in resultats if x.get("societe") == nom_soc), None)
        if ex:
            row["siren"] = ex.get("siren") or row.get("siren", "")
            row["domaine"] = ex.get("domaine") or row.get("domaine", "")
        try:
            data = await _enrich_claude_core({**row, "max_contacts": max_ajouter})
            noms_existants = [f"{x.get('prenom','')} {x.get('nom_dg','')}".lower().strip()
                              for x in resultats if x.get("societe") == nom_soc]
            ajoutes = 0
            for ct in data.get("contacts", []):
                if ajoutes >= max_ajouter:
                    break
                key = f"{ct.get('prenom','')} {ct.get('nom','')}".lower().strip()
                if key not in noms_existants:
                    resultats.append({
                        "org_id": row.get("org_id", ""), "societe": nom_soc,
                        "siren": row.get("siren", ""), "domaine": row.get("domaine", ""),
                        "adresse": row.get("adresse", ""),
                        "prenom": ct.get("prenom", ""), "nom_dg": ct.get("nom", ""),
                        "titre": ct.get("titre", ""), "email": ct.get("email", "") or "",
                        "confiance": ct.get("confiance_email", "faible"),
                        "source": "Claude+web", "linkedin": "", "dans_pipedrive": "",
                    })
                    ajoutes += 1
        except Exception as e:
            print(f"[RUN {r['id']}] phase2 erreur sur {nom_soc}: {e}")
        r["progres_traites"] = i + 1
        _persist_run(r["id"])
        if i < len(societes) - 1 and not r["stop_demande"]:
            await asyncio.sleep(_DELAI_SOCIETE)


# ─── PHASE 3 : Pipedrive — emails des contacts déjà au CRM ──────────
async def _run_phase3(r: dict):
    r["phase_courante"] = "3. Pipedrive"
    resultats = r["resultats"]
    sans = [x for x in resultats if x.get("nom_dg") and not x.get("email")]
    r["progres_total"] = len(sans)
    r["progres_traites"] = 0
    _persist_run(r["id"])
    for i, x in enumerate(sans):
        if r["stop_demande"]:
            return
        try:
            pd = await check_pipedrive(x.get("prenom", ""), x.get("nom_dg", ""))
            if pd:
                # Politique 09/06/2026 : on ne récupère plus l'email Pipedrive,
                # uniquement le téléphone et le poste. Kaspr/FullEnrich
                # (phases 4 et 5) se chargent d'aller chercher un email frais.
                if pd.get("phone") and not x.get("phone"):
                    x["phone"] = pd["phone"]
                if pd.get("job_title"):
                    x["titre"] = pd["job_title"]   # le poste du CRM prime
                x["dans_pipedrive"] = "oui"
                src = (x.get("source", "") or "")
                if "Pipedrive" not in src:
                    x["source"] = (src + "+Pipedrive").lstrip("+")
        except Exception as e:
            print(f"[RUN {r['id']}] phase3 erreur: {e}")
        r["progres_traites"] = i + 1
        _persist_run(r["id"])


# ─── PHASES 4 & 5 : Kaspr / FullEnrich — enrichissement emails ──────
# phase 'kaspr'      → _enrich_emails_core ne fait QUE Kaspr.
# phase 'fullenrich' → _enrich_emails_core ne fait QUE FullEnrich.
# Les deux phases sont donc bien distinctes (le pas-à-pas est honnête).
# Le FILTRE des contacts envoyés diffère aussi entre 4 et 5 (cf JS runEmailBatch).
async def _run_phase_emails(r: dict, phase: str):
    label = "Kaspr" if phase == "kaspr" else "Fullenrich"
    r["phase_courante"] = f"{'4' if phase == 'kaspr' else '5'}. {label}"
    resultats = r["resultats"]
    to_enrich = []
    for idx, x in enumerate(resultats):
        if not x.get("nom_dg"):
            continue
        # FullEnrich n'est sollicité que pour l'email (décision Marie 08/06/2026)
        # → on n'envoie plus les contacts qui ont déjà un email même sans phone.
        sans_email = (not x.get("email")) or x.get("confiance") == "faible"
        if sans_email:
            to_enrich.append({**x, "idx": idx})
    if not to_enrich:
        r["progres_total"] = 0
        r["progres_traites"] = 0
        _persist_run(r["id"])
        return
    total_batches = (len(to_enrich) + _BATCH_SIZE - 1) // _BATCH_SIZE
    r["progres_total"] = total_batches
    r["progres_traites"] = 0
    _persist_run(r["id"])
    for b in range(total_batches):
        if r["stop_demande"]:
            return
        batch = to_enrich[b * _BATCH_SIZE:(b + 1) * _BATCH_SIZE]
        try:
            data = await _enrich_emails_core({
                "phase": phase,
                "_stop_check": lambda: r["stop_demande"],
                "contacts": [{
                    "idx": x["idx"], "prenom": x.get("prenom", ""),
                    "nom": x.get("nom_dg", ""), "domaine": x.get("domaine", ""),
                    "linkedin": x.get("linkedin", ""),
                    "societe": x.get("societe", ""), "siren": x.get("siren", ""),
                    "email": x.get("email", ""), "confiance": x.get("confiance", ""),
                } for x in batch],
            })
            emails = data.get("emails", {})
            for idx_str, val in emails.items():
                idx = int(idx_str)
                if not (0 <= idx < len(resultats)):
                    continue
                if isinstance(val, str):
                    email, linkedin, phone, srctag = val, "", "", f"+{label}"
                else:
                    email    = val.get("email", "") if val else ""
                    linkedin = val.get("linkedin", "") if val else ""
                    phone    = val.get("phone", "") if val else ""
                    srctag   = val.get("source", f"+{label}") if val else f"+{label}"
                if phone and not resultats[idx].get("phone"):
                    resultats[idx]["phone"] = phone
                if email and "@" in email:
                    resultats[idx]["email"] = email
                    resultats[idx]["confiance"] = "haute"
                    if linkedin:
                        resultats[idx]["linkedin"] = linkedin
                    src = resultats[idx].get("source", "") or ""
                    tag = srctag.replace("+", "")
                    if tag not in src:
                        resultats[idx]["source"] = src + srctag
                elif linkedin and not resultats[idx].get("linkedin"):
                    resultats[idx]["linkedin"] = linkedin
        except Exception as e:
            print(f"[RUN {r['id']}] {label} lot {b+1}/{total_batches} échoué: {e}")
        r["progres_traites"] = b + 1
        _persist_run(r["id"])
        if b < total_batches - 1 and not r["stop_demande"]:
            await asyncio.sleep(_PAUSE_BATCH)


# ─── Finalisation : Excel + email + archivage page /runs ────────────
def _envoyer_excel_email(emails_dest: list, rows: list, filename: str):
    """Envoi SMTP bloquant (à appeler via asyncio.to_thread)."""
    excel_content = generer_excel([dict(x) for x in rows])
    msg = MIMEMultipart()
    msg['From']    = SMTP_USER
    msg['To']      = ", ".join(emails_dest)
    emails_count   = len([x for x in rows if x.get('email')])
    msg['Subject'] = f"Enrichissement dirigeants — {len(rows)} contacts"
    body = (f"Bonjour,\n\nVotre enrichissement est terminé.\n"
            f"{len(rows)} contacts exportés dont {emails_count} emails trouvés.\n"
            f"Fichier Excel en pièce jointe.\n\nEnrichisseur Dirigeants")
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    part.set_payload(excel_content)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
    msg.attach(part)
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, emails_dest, msg.as_string())
    print(f"[EMAIL] ✅ Excel envoyé à {', '.join(emails_dest)}")


async def _finaliser_run(r: dict):
    """En fin de run : envoie l'Excel par email (si destinataires) et
    archive le run dans runs_history (pour la page /runs)."""
    rows = r["resultats"]
    if not rows:
        return
    fname = r["excel_filename"] or f"enrichissement_{r['id']}.xlsx"
    emails_dest = [e for e in r["emails_dest"] if e and "@" in e]
    if emails_dest and SMTP_USER and SMTP_PASS:
        try:
            await asyncio.to_thread(_envoyer_excel_email, emails_dest, rows, fname)
        except Exception as e:
            print(f"[RUN {r['id']}] envoi email échoué: {e}")
    # Archive dans runs_history → la page /runs continue de fonctionner
    try:
        emails_count = len([x for x in rows if x.get("email")])
        nb_phones    = sum(1 for x in rows if x.get("phone"))
        with _sqlite_conn() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO runs_history (id, created_at, filename, "
                "total_contacts, emails_count, phones_count, recipient, rows_json) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (r["id"], int(time.time()), fname, len(rows), emails_count,
                 nb_phones, ", ".join(emails_dest),
                 json.dumps(rows, ensure_ascii=False)))
            conn.commit()
        print(f"[RUN {r['id']}] Archivé dans runs_history")
    except Exception as e:
        print(f"[RUN {r['id']}] archive runs_history échouée: {e}")


# ════════════════════════════════════════════════════════════════════
# ENDPOINTS /run/* — pilotage des runs depuis le frontend
# ════════════════════════════════════════════════════════════════════
def _run_public(r: dict, light: bool = True) -> dict:
    """Vue JSON d'un run. light=True → sans entree/resultats (pour /run/list)."""
    d = {
        "id": r["id"], "nom": r["nom"], "mode": r["mode"], "phases": r["phases"],
        "statut": r["statut"], "phase_courante": r["phase_courante"],
        "progres_traites": r["progres_traites"], "progres_total": r["progres_total"],
        "created_at": r["created_at"], "started_at": r["started_at"],
        "finished_at": r["finished_at"], "erreur": r["erreur"],
        "rang": _rang_file(r),
        "nb_contacts": len(r["resultats"]),
        "nb_emails": len([x for x in r["resultats"] if x.get("email")]),
        "nb_telephones": len([x for x in r["resultats"] if x.get("phone")]),
    }
    if not light:
        d["resultats"]   = r["resultats"]
        d["emails_dest"] = r["emails_dest"]
    return d


@app.post("/run/create")
async def run_create(request: Request):
    """Crée un run et le met EN_ATTENTE. Le worker le prendra en charge.
    Payload : {nom, mode:'societes'|'contacts', phases:[1..5],
               rows:[...] (mode societes) | contacts:[...] (mode contacts),
               emails_dest:[...] ou "a@b.com, c@d.com"}"""
    data = await request.json()
    mode = data.get("mode", "societes")
    if mode not in ("societes", "contacts"):
        mode = "societes"
    # Phases demandées (le pas-à-pas : l'utilisateur peut n'en cocher que certaines)
    phases = data.get("phases") or ([4, 5] if mode == "contacts" else [1, 2, 3, 4, 5])
    phases = sorted({int(p) for p in phases if int(p) in (1, 2, 3, 4, 5)})
    if mode == "contacts":
        # Pas de Pappers/Claude sur des contacts déjà identifiés
        phases = [p for p in phases if p in (3, 4, 5)] or [4, 5]
    if not phases:
        return JSONResponse({"error": "Aucune phase valide demandée"}, status_code=400)
    entree = (data.get("contacts") if mode == "contacts" else data.get("rows")) or []
    if not entree:
        return JSONResponse({"error": "Aucune donnée en entrée"}, status_code=400)
    emails_dest = data.get("emails_dest", []) or []
    if isinstance(emails_dest, str):
        emails_dest = emails_dest.split(",")
    emails_dest = [e.strip() for e in emails_dest if e and "@" in e]
    run_id = _uuid.uuid4().hex[:12]
    nom_brut = (data.get("nom") or "").strip()
    nom = nom_brut or f"run du {datetime.now().strftime('%d/%m %Hh%M')}"
    safe = re.sub(r'[^a-zA-Z0-9_\-]', '_', nom_brut or "enrichissement")
    today = datetime.now().strftime("%Y-%m-%d")
    r = {
        "id": run_id, "nom": nom, "mode": mode, "phases": phases,
        "statut": "EN_ATTENTE", "phase_courante": "", "progres_traites": 0,
        "progres_total": len(entree), "created_at": int(time.time()),
        "started_at": None, "finished_at": None, "entree": entree,
        "resultats": [], "emails_dest": emails_dest, "erreur": "",
        "stop_demande": False,
        "excel_filename": f"enrichissement_{safe}_{today}.xlsx",
    }
    _RUNS[run_id] = r
    _persist_run(run_id)
    rang = _rang_file(r)
    print(f"[RUN {run_id}] Créé — {nom} | mode={mode} | phases={phases} | rang file={rang}")
    return {"ok": True, "run_id": run_id, "rang": rang}


@app.get("/run/list")
async def run_list():
    """Liste les 50 runs les plus récents (vue légère, pour le polling)."""
    runs = sorted(_RUNS.values(), key=lambda r: r["created_at"], reverse=True)
    return {"runs": [_run_public(r, light=True) for r in runs[:50]]}


@app.get("/run/{run_id}")
async def run_get(run_id: str):
    """État complet d'un run, résultats inclus (pour réafficher le tableau)."""
    r = _RUNS.get(run_id)
    if not r:
        return JSONResponse({"error": "Run introuvable"}, status_code=404)
    return _run_public(r, light=False)


@app.post("/run/{run_id}/stop")
async def run_stop(run_id: str):
    """Demande l'arrêt d'un run. EN_ATTENTE → annulé direct. EN_COURS →
    le worker s'arrête proprement à la prochaine vérification."""
    r = _RUNS.get(run_id)
    if not r:
        return JSONResponse({"error": "Run introuvable"}, status_code=404)
    if r["statut"] == "EN_ATTENTE":
        r["statut"] = "ARRETE"
        r["phase_courante"] = "Annulé avant démarrage"
        r["finished_at"] = int(time.time())
    elif r["statut"] == "EN_COURS":
        r["stop_demande"] = True
    _persist_run(run_id)
    return {"ok": True, "statut": r["statut"]}


@app.get("/run/{run_id}/excel")
async def run_excel(run_id: str):
    """Télécharge l'Excel d'un run (régénéré depuis ses résultats)."""
    r = _RUNS.get(run_id)
    if not r or not r["resultats"]:
        return JSONResponse({"error": "Run introuvable ou sans résultats"}, status_code=404)
    content = generer_excel([dict(x) for x in r["resultats"]])
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{r["excel_filename"]}"'})


@app.post("/run/{run_id}/continue")
async def run_continue(run_id: str, request: Request):
    """Pas-à-pas serveur : relance un run terminé/arrêté avec d'autres
    phases, en repartant de ses résultats actuels (ex : lancer Kaspr+
    FullEnrich après n'avoir fait que Pappers+Claude)."""
    r = _RUNS.get(run_id)
    if not r:
        return JSONResponse({"error": "Run introuvable"}, status_code=404)
    if r["statut"] not in ("TERMINE", "ERREUR", "ARRETE"):
        return JSONResponse({"error": "Run encore actif — attendez la fin"}, status_code=400)
    data = await request.json()
    phases = sorted({int(p) for p in (data.get("phases") or []) if int(p) in (1, 2, 3, 4, 5)})
    if not phases:
        return JSONResponse({"error": "Aucune phase valide demandée"}, status_code=400)
    r["phases"] = phases
    r["statut"] = "EN_ATTENTE"
    r["phase_courante"] = ""
    r["progres_traites"] = 0
    r["progres_total"] = 0
    r["erreur"] = ""
    r["stop_demande"] = False
    r["finished_at"] = None
    r["created_at"] = int(time.time())   # repart en fin de file
    _persist_run(run_id)
    print(f"[RUN {run_id}] ↻ Relancé (pas-à-pas) avec phases {phases}")
    return {"ok": True, "run_id": run_id, "rang": _rang_file(r)}
